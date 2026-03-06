import pandas as pd
import yfinance as yf
import numpy as np
import os
import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import importlib
import holidays


# ALLGEMEINE FUNKTIONEN

# ahlib einbinden
try:
    from ahlib import (
        create_extended_logger,
        ExtendedLogger,
        export_2D_df_to_excel_clean_table,
        format_excel_as_table_with_freeze,
        format_excel_columns,
        set_working_directory,
        settings_import,
        files_availability_check,
        export_df_to_excel,
        import_parquet
    )
    print("Import der ahlib Bibliothek erfolgreich")
except ImportError as e:
    sys.exit(f"Fehler beim Import der ahlib Bibliothek: {e}")


# Umrechnung eines 2D Dataframes mit Datum als Index in die Prozentwerte "Spaltenwert / Summe über alle Spalten"
def df_transform_each_line_to_percentage(df):
    """
    Berechnet für jedes Datum und jeden Typ den prozentualen Anteil des Wertes einer Spalte im Verhältnis zur
    Gesamtsumme aller Typen an diesem Datum. Der Spaltenname wird automatisch bestimmt.
    
    Parameter:
        df (DataFrame): Ein DataFrame mit MultiIndex (date, type) und einer einzigen Wertspalte.

    Rückgabe:
        DataFrame: Ein DataFrame mit denselben Indizes und Spalte '<column_name>_percentage',
                   die den prozentualen Anteil angibt.
    """
    # Bestimme den Namen der Wertspalte
    value_column = df.columns[0]
    
    # Berechne die Summe der Werte für jeden Tag
    daily_totals = df.groupby(level='date')[value_column].transform('sum')
    
    # Berechne den Prozentsatz für jeden Typ pro Datum
    df[f'{value_column}_percentage'] = (df[value_column] / daily_totals)
    
    return df[[f'{value_column}_percentage']]

# Reduktion eines Dataframes durch Selektion des Monatsenddatums bzw. Datums des letzten Tags im laufenden Monat
def df_to_eom(df):
    """
    Reduziert den DataFrame auf die Einträge des jeweils jüngsten Datums pro Monat.
    Es werden nur Einträge mit dem entsprechenden Datum selektiert. Es erfolgt keine Operation der Daten
    Falls das größte Datum des DataFrames nicht in der Liste latest_dates_per_month enthalten ist,
    wird es zur Liste hinzugefügt.

    Parameter:
        df (DataFrame): Ein DataFrame mit einem MultiIndex, der einen 'date'-Index enthält.
                        Wie der Multiindex aufgebaut ist, ist unwichtig, solange 'date' entahalten ist

    Rückgabe:
        DataFrame: Der reduzierte DataFrame mit nur den Einträgen des jeweils jüngsten Datums pro Monat.
    """
    # Bestimme das jeweils größte Datum pro Monat
    latest_dates_per_month = df.index.get_level_values('date').to_period('M').drop_duplicates().to_timestamp('M')
    
    # Bestimme das jüngste (größte) Datum im DataFrame
    latest_date_in_df = df.index.get_level_values('date').max()
    
    # Falls das jüngste Datum des DataFrames nicht in latest_dates_per_month enthalten ist, ergänze es
    if latest_date_in_df not in latest_dates_per_month:
        latest_dates_per_month = latest_dates_per_month.append(pd.Index([latest_date_in_df]))
    
    # Filtere den DataFrame basierend auf den bestimmten Monatsenddaten
    df_eom = df[df.index.get_level_values('date').isin(latest_dates_per_month)]
    
    return df_eom

def df_to_eoy(df):
    """
    Reduziert den DataFrame auf die Einträge des jeweils jüngsten Datums pro Jahr.
    Es werden nur Einträge mit dem entsprechenden Datum selektiert. Es erfolgt keine Operation der Daten.
    Falls das größte Datum des DataFrames nicht in der Liste latest_dates_per_year enthalten ist,
    wird es zur Liste hinzugefügt.

    Parameter:
        df (DataFrame): Ein DataFrame mit einem MultiIndex, der einen 'date'-Index enthält.

    Rückgabe:
        DataFrame: Der reduzierte DataFrame mit nur den Einträgen des jeweils jüngsten Datums pro Jahr.
    """
    # Bestimme das jeweils größte Datum pro Jahr
    latest_dates_per_year = df.index.get_level_values('date').to_period('Y').drop_duplicates().to_timestamp('Y')

    # Bestimme das jüngste (größte) Datum im DataFrame
    latest_date_in_df = df.index.get_level_values('date').max()

    # Falls das jüngste Datum des DataFrames nicht in latest_dates_per_year enthalten ist, ergänze es
    if latest_date_in_df not in latest_dates_per_year:
        latest_dates_per_year = latest_dates_per_year.append(pd.Index([latest_date_in_df]))

    # Filtere den DataFrame basierend auf den bestimmten Jahresenddaten
    df_eoy = df[df.index.get_level_values('date').isin(latest_dates_per_year)]

    return df_eoy

def df_2D_sum_per_period(df_input, period='month'):
    """
    Aggregiert eine numerische Spalte eines DataFrames mit MultiIndex (Datum, ID) über einen Zeitraum ('month' oder 'year').
    Das Ergebnis erhält das jeweils jüngste Datum im Zeitraum als neuen Indexwert.

    Voraussetzungen:
    - Der DataFrame muss einen MultiIndex mit zwei Ebenen haben.
    - Eine Ebene muss datetime-basiert sein, die andere eine ID (z. B. 'wkn').
    - Der DataFrame muss genau eine numerische Spalte zur Aggregation enthalten.

    Parameter:
        df_input (DataFrame): Eingabedaten mit MultiIndex (datetime, id).
        period (str): 'month' oder 'year'

    Rückgabe:
        DataFrame mit MultiIndex (date, id) und aggregierter Spalte.
    """
    import pandas as pd

    # Gültige Perioden
    valid_periods = ['month', 'year']
    if period not in valid_periods:
        raise ValueError(f"Ungültiger Zeitraum: '{period}'. Gültige Optionen sind {valid_periods}.")

    # Sicherheitskopie
    df = df_input.copy()

    # Prüfe MultiIndex-Struktur
    if not isinstance(df.index, pd.MultiIndex) or len(df.index.names) != 2:
        raise ValueError("DataFrame muss einen MultiIndex mit genau zwei Ebenen haben.")

    # Index-Ebenen dynamisch identifizieren
    level_0 = df.index.get_level_values(0)
    level_1 = df.index.get_level_values(1)

    if pd.api.types.is_datetime64_any_dtype(level_0) and not pd.api.types.is_datetime64_any_dtype(level_1):
        date_level = level_0
        id_level = level_1
        date_name = df.index.names[0]
        id_name = df.index.names[1]
    elif pd.api.types.is_datetime64_any_dtype(level_1) and not pd.api.types.is_datetime64_any_dtype(level_0):
        date_level = level_1
        id_level = level_0
        date_name = df.index.names[1]
        id_name = df.index.names[0]
    else:
        raise ValueError("Der MultiIndex muss eine Datumsebene und eine ID-Ebene enthalten.")

    # Aggregationsspalte erkennen (es darf nur eine geben)
    value_columns = [col for col in df.columns if pd.api.types.is_numeric_dtype(df[col])]
    if len(value_columns) != 1:
        raise ValueError("Es muss genau eine numerische Spalte zur Aggregation vorhanden sein.")
    value_col = value_columns[0]

    # Jahr (und ggf. Monat) extrahieren
    df['date_only'] = date_level
    df['id_value'] = id_level
    df['year'] = df['date_only'].dt.year

    if period == 'month':
        df['month'] = df['date_only'].dt.month
        group_levels = ['year', 'month', 'id_value']
        period_levels = ['year', 'month']
    else:
        group_levels = ['year', 'id_value']
        period_levels = ['year']

    # Aggregieren
    aggregated_sum = df.groupby(group_levels)[value_col].sum().reset_index()

    # Letztes Datum der Periode bestimmen
    latest_dates = df.groupby(period_levels)['date_only'].max().reset_index()

    # Zusammenführen
    merged = pd.merge(aggregated_sum, latest_dates, on=period_levels, how='left')
    merged.rename(columns={'date_only': 'date'}, inplace=True)

    # MultiIndex setzen
    merged.set_index(['date', 'id_value'], inplace=True)
    merged.index.set_names(['date', id_name], inplace=True)
    merged.sort_index(inplace=True)

    return merged[[value_col]]

def df_1D_sum_per_period(df_input, period='month'):
    """
    Aggregiert eine Zeitreihe mit DatetimeIndex über definierte Zeiträume ('month' oder 'year').
    Erwartet einen DataFrame mit genau einer numerischen Spalte und einem DatetimeIndex.

    Parameter:
        df_input (DataFrame): DataFrame mit DatetimeIndex und einer Spalte (z. B. 'invest').
        period (str): Aggregationszeitraum ('month' oder 'year').

    Rückgabe:
        DataFrame mit DatetimeIndex (jeweils jüngstes Datum des Zeitraums) und aggregierter Wertespalte.
    """
    import pandas as pd

    valid_periods = ['month', 'year']
    if period not in valid_periods:
        raise ValueError(f"Ungültiger Zeitraum: '{period}'. Gültige Optionen sind {valid_periods}.")

    # Sicherheitskopie
    df = df_input.copy()

    # Prüfungen
    if not isinstance(df.index, pd.DatetimeIndex):
        raise ValueError("Der Index muss ein DatetimeIndex sein.")
    if len(df.columns) != 1:
        raise ValueError("Der DataFrame muss genau eine Spalte enthalten.")

    value_col = df.columns[0]

    # Zeitdimension vorbereiten
    df['date_only'] = df.index
    df['year'] = df['date_only'].dt.year
    if period == 'month':
        df['month'] = df['date_only'].dt.month
        group_keys = ['year', 'month']
    else:
        group_keys = ['year']

    # Werte aggregieren
    aggregated = df.groupby(group_keys)[value_col].sum().reset_index()

    # Jüngstes Datum je Zeitraum ermitteln
    latest_dates = df.groupby(group_keys)['date_only'].max().reset_index()

    # Zusammenführen
    merged = pd.merge(aggregated, latest_dates, on=group_keys, how='left')
    merged.rename(columns={'date_only': 'date'}, inplace=True)

    # Index setzen und zurückgeben
    merged.set_index('date', inplace=True)
    merged.sort_index(inplace=True)

    return merged[[value_col]]

def function_result(function_name, error_count, warning_count, logger):
    """
    Überprüft die Rückgabewerte einer Funktion und gibt eine entsprechende Meldung aus.
    Beendet das Programm, wenn Fehler aufgetreten sind.

    Parameter:
        function_name (str): Der Name der aufgerufenen Funktion.
        error_count (int): Anzahl der aufgetretenen Fehler.
        warning_count (int): Anzahl der aufgetretenen Warnungen.
        logger (ExtendedLogger): Logger instance.
    """
    # Überprüfung der Rückgabewerte für Fehler und Warnugen
    if error_count > 0:
        logger.error(f"{function_name} fehlgeschlagen. Das Programm wird beendet.")
        sys.exit(1)
    elif warning_count > 0:
        logger.warning(f"{function_name} abgeschlossen mit {warning_count} Warnung(en).")
    else:
        logger.info(f"{function_name} erfolgreich abgeschlossen.")

def export_2D_df_to_excel_format(df, export, logger):
    """
    Exportiert einen datframe mit 2D Multindex nach Excel als Tablle mit Formaten

    erwartet in export ein Dicutionary mit enabled, filename, list von format strings, liste von spalten breiten

    """

    enabled = export.get("enabled", False)
    filename = export.get("filename", "")
    format_numbers = export.get("column_formats", {})
    format_columns = export.get("column_widths", {})

    if enabled:
        export_2D_df_to_excel_clean_table(df, filename, logger)
        format_excel_as_table_with_freeze(filename, logger, table_name="Table1", style_name="TableStyleMedium1", freeze_first_row=True)
        format_excel_columns(filename, format_numbers, logger, format_columns)

    return





# Spezifische Funktionen
# Funktion die aus dem instruments file (EXCEL) die Schlüssel wkn, ticker, Name und Default Wert lädet
def instruments_import(filename, logger):
    """
    Liest die Excel-Datei und importiert die ersten vier Spalten (wkn, ticker, instrument_name, Default)
    in einen Pandas DataFrame. wkn und ticker werden in Kleinbuchstaben umgewandelt.
    wkn wird als Index gesetzt. Spaltennamen werden auf 'ticker', 'Name', 'default_value' gesetzt.

    Fehlerabfrage: Wenn die Datei kein Excel-Format hat oder ein anderer Fehler auftritt, wird eine Meldung ausgegeben.
    """
    try:
        # Prüfe, ob die Datei eine Excel-Datei ist
        if not filename.endswith(('.xlsx', '.xls')):
            raise ValueError(f"Die Datei '{filename}' ist keine Excel-Datei.")

        # Lese die ersten vier Spalten aus der Excel-Datei und setze die erste Spalte (wkn) als Index
        df = pd.read_excel(filename, usecols=[0, 1, 2, 3], index_col=0)

        # Wandle den Index (wkn) und die ticker-Spalte in Kleinbuchstaben um
        df.index = df.index.str.lower()  # wkn auf Kleinbuchstaben umstellen
        df['ticker'] = df['ticker'].str.lower()  # ticker auf Kleinbuchstaben umstellen

        # Setze die Spaltennamen sicher
        df.columns = ['ticker', 'instrument_name', 'default_value']  # Sichere Zuweisung der Spaltennamen

        # Gib den DataFrame zurück
        return df

    except FileNotFoundError:
        logger.error(f"Die Datei '{filename}' wurde nicht gefunden.")
        return None
    except ValueError as ve:
        logger.error(f"{ve}")
        return None
    except Exception as e:
        logger.error(f"Ein Fehler ist aufgetreten: {e}")
        return None

# Funktion die aus dem instruments file (EXCEL) die Spalten mit relativen Anteilen für Detailanalysen lädet
def instruments_details_import(filename, search_param, logger):
    """
    Liest die Excel-Datei und importiert die 'wkn'-Spalte und alle Spalten, deren Namen mit dem Suchparameter beginnen,
    in einen Pandas DataFrame. Der MultiIndex besteht aus 'wkn' und dem in Kleinbuchstaben übergebenen Suchparameter,
    wobei die Spalte 'share' die Anteile enthält. Fehlende Werte werden auf 0 gesetzt, und alle Index- und Spaltennamen
    werden in Kleinbuchstaben umgewandelt.

    Parameter:
        filename (str): Pfad zur Excel-Datei.
        search_param (str): Präfix für die Spaltennamen (z. B. 'Reg'), die importiert werden sollen.

    Rückgabe:
        DataFrame: Ein DataFrame mit MultiIndex (wkn, <search_param>) und der Spalte 'share'.
    """
    try:
        # Prüfe, ob die Datei eine Excel-Datei ist
        if not filename.endswith(('.xlsx', '.xls')):
            raise ValueError(f"Die Datei '{filename}' ist keine Excel-Datei.")

        # Lese die Excel-Datei ein
        df = pd.read_excel(filename)

        # Wähle nur die Spalten aus, die mit dem Suchparameter beginnen, sowie die 'wkn'-Spalte
        relevant_columns = ['wkn'] + [col for col in df.columns if col.startswith(search_param)]
        df = df[relevant_columns]

        # Setze 'wkn' als Index und wandle ihn in Kleinbuchstaben um
        df['wkn'] = df['wkn'].str.lower()
        df.set_index('wkn', inplace=True)

        # Setze alle fehlenden Werte auf 0
        df.fillna(0, inplace=True)

        # Erstelle neue Spaltennamen ohne das Präfix und wandle sie in Kleinbuchstaben um
        search_param_lower = search_param.lower()
        df.columns = [col.lower().replace(f"{search_param} ", "") for col in df.columns]

        # Staple die Spalten in Zeilen, um einen MultiIndex mit 'wkn' und '<search_param>' zu erzeugen
        df = df.stack().reset_index()
        df.columns = ['wkn', search_param_lower, 'share']

        # Setze den MultiIndex (wkn, <search_param>) und wandle alles in Kleinbuchstaben um
        df.set_index(['wkn', search_param_lower], inplace=True)
        df.index.names = [name.lower() for name in df.index.names]
        df.columns = [col.lower() for col in df.columns]

        return df

    except FileNotFoundError:
        logger.error(f"Die Datei '{filename}' wurde nicht gefunden.")
        return None
    except ValueError as ve:
        logger.error(f"{ve}")
        return None
    except Exception as e:
        logger.error(f"Ein Fehler ist aufgetreten: {e}")
        return None
       
# Funktion zum gegenseitigen Abgleich ob alle wkn in prices und instrumentes enthalten sind
def prices_check_for_instruments(prices, instruments, logger):
    """
    Überprüft, ob alle WKNs aus 'prices' im DataFrame 'instruments' vorhanden sind und umgekehrt.
    
    Parameter:
        prices (DataFrame): Der DataFrame mit Preis-Daten.
        instruments (DataFrame): Der DataFrame mit Instruments-Daten.
        logfile (str): Der Name des Logfiles.
        screen (bool): Wenn True, werden Bildschirmmeldungen angezeigt.
        log (bool): Wenn True, werden Meldungen ins Logfile geschrieben.
    """
    wkn_prices = set(prices.index.get_level_values('wkn'))
    wkn_instruments = set(instruments.index)

    missing_in_instruments = wkn_prices - wkn_instruments
    if missing_in_instruments:
        logger.warning(f"Die folgenden WKNs aus 'prices' fehlen in 'instruments': {missing_in_instruments}")
    else:
        logger.info("Alle WKNs aus 'prices' sind in 'instruments' vorhanden.")

    missing_in_prices = wkn_instruments - wkn_prices
    if missing_in_prices:
        logger.warning(f"Die folgenden WKNs aus 'instruments' fehlen in 'prices': {missing_in_prices}")
    else:
        logger.info("Alle WKNs aus 'instruments' sind in 'prices' vorhanden.")

# Funktion zum Aktualisieren der Kursdaten zwischen last_date und gestern
def prices_update(prices, instruments, logger):
    """
    Aktualisiert fehlende Kursdaten in 'prices' zwischen dem letzten Datum und gestern,
    wobei nur Handelstage (Mo–Fr, ohne Feiertage) berücksichtigt werden.
    
    Parameter:
        prices (DataFrame): Bestehende Kursdaten. MultiIndex ('date', 'wkn'), Spalte 'price'
        instruments (DataFrame): Enthält je WKN einen 'ticker' und optional 'default_value'
        logger (ExtendedLogger): Logger instance for output and logging.

    Rückgabe:
        DataFrame: Aktualisierter 'prices'-DataFrame
    """
    # Aktuelles Datum (nur ohne Uhrzeit)
    today = datetime.today().date()
    yesterday = pd.Timestamp(today - timedelta(days=1))

    # Letztes verfügbares Datum im DataFrame
    last_date = prices.index.get_level_values('date').max()

    # Deutsche Feiertage
    de_holidays = holidays.Germany()

    # Datumsbereich: alle Kalendertage zwischen letztem Kursdatum und gestern
    all_dates = pd.date_range(start=last_date + timedelta(days=1), end=yesterday)

    # Nur Mo–Fr und keine Feiertage
    missing_dates = [d for d in all_dates if d.weekday() < 5 and d.strftime('%Y-%m-%d') not in de_holidays]

    if not missing_dates:
        logger.info(f"Keine fehlenden Handelstage zwischen {last_date.date()} und {yesterday.date()}")
        return prices

    # Für jede WKN einzeln Kursdaten abrufen
    for wkn, row in instruments.iterrows():
        #ticker = row['ticker']
        raw_ticker = row['ticker']

        # Erst prüfen, ob NaN oder leer
        if pd.isna(raw_ticker) or str(raw_ticker).strip() == '':
            ticker = None
        else:
            ticker = str(raw_ticker).strip().upper()

        
        default_value = row['default_value']
        #print(wkn, "Defalut Value", default_value) #debug
        

        if pd.notna(ticker) and ticker.strip() != '':
            try:
                data = yf.download(
                    ticker,
                    start=missing_dates[0],
                    end=missing_dates[-1] + timedelta(days=1),  # Enddatum exklusiv
                    progress=False,
                    auto_adjust=False
                )

                if data.empty:
                    logger.warning(
                        f"Keine Daten für Ticker {ticker} im Zeitraum {missing_dates[0].date()} bis {missing_dates[-1].date()}"
                    )
                    continue

                # Stelle sicher, dass der Index normiert ist
                data.index = data.index.normalize()

                for date in missing_dates:
                    normalized_date = pd.Timestamp(date).normalize()

                    try:
                        # Versuche Zugriff auf 'Close'-Wert
                        close_entry = data.loc[normalized_date, 'Close']

                        # Falls Series: z. B. durch mehrdimensionale Struktur
                        if isinstance(close_entry, pd.Series):
                            close_value = close_entry.iloc[0]
                        else:
                            close_value = close_entry

                        prices.loc[(normalized_date, wkn), 'price'] = close_value

                    except KeyError:
                        prices.loc[(normalized_date, wkn), 'price'] = np.nan
                        logger.warning(
                            f"Kein Kurs für {ticker} am {normalized_date.date()} verfügbar"
                        )

                    except Exception as e:
                        logger.error(
                            f"Unerwarteter Fehler bei Zugriff auf Close-Wert {ticker} am {normalized_date.date()}: {e}"
                        )

            except Exception as e:
                logger.error(
                    f"Fehler beim Abrufen der Daten am {missing_dates[0]} für WKN {wkn} ({ticker}): {e}"
                )

        else:
            # Kein Ticker → Defaultwert setzen
            for date in missing_dates:
                normalized_date = pd.Timestamp(date).normalize()
                prices.loc[(normalized_date, wkn), 'price'] = float(default_value) if default_value is not None else np.nan

    return prices

# Funktion zum Einlesen der Buchungsdaten
def bookings_import(filename, logger):
    try:
        # Lese die ersten vier Spalten aus der Excel-Datei
        df = pd.read_excel(filename, usecols=[0, 1, 2, 3], names=['date', 'wkn', 'bank', 'delta'])

        # Konvertiere 'wkn' und 'bank' in Kleinbuchstaben
        df['wkn'] = df['wkn'].str.lower()
        df['bank'] = df['bank'].str.lower()

        # Entferne Zeilen mit NaN in 'wkn', 'bank' oder 'delta'
        df.dropna(subset=['wkn', 'bank', 'delta'], inplace=True)

        # Setze den MultiIndex auf 'date', 'wkn', 'bank'
        df.set_index(['date', 'wkn', 'bank'], inplace=True)

        # Fasse Einträge mit demselben MultiIndex zusammen und summiere 'delta'
        # Damit werden mehrere Transaktion an einem Tag für eine WKN (bei der gleichen bank) zu einem Eintrag kombiniert
        df = df.groupby(level=['date', 'wkn', 'bank']).sum()

        return df

    except FileNotFoundError:
        logger.error(f"Die Datei '{filename}' wurde nicht gefunden.")
        return None
    except Exception as e:
        logger.error(f"Ein Fehler ist beim Import der Buchungen aus '{filename}' aufgetreten: {e}")
        return None
    
def fees_import(filename, logger):
    try:
        # Lese die ersten vier Spalten aus der Excel-Datei
        df = pd.read_excel(filename, usecols=[0, 1, 2, 5], names=['date', 'wkn', 'bank', 'delta', 'invest_divest', 'fees'])

        # Konvertiere 'wkn' und 'bank' in Kleinbuchstaben
        df['wkn'] = df['wkn'].str.lower()
        df['bank'] = df['bank'].str.lower()

        # Entferne Zeilen mit NaN in 'wkn', 'bank' oder 'fees'
        df.dropna(subset=['wkn', 'bank', 'fees'], inplace=True)

        # Setze den MultiIndex auf 'date', 'wkn', 'bank'
        df.set_index(['date', 'wkn', 'bank'], inplace=True)

        # Fasse Einträge mit demselben MultiIndex zusammen und summiere die Wert-Spalte
        # Damit werden mehrere Transaktion an einem Tag für eine WKN (bei der gleichen bank) zu einem Eintrag kombiniert
        df = df.groupby(level=['date', 'wkn', 'bank']).sum()

        return df

    except FileNotFoundError:
        logger.error(f"Die Datei '{filename}' wurde nicht gefunden.")
        return None
    except Exception as e:
        logger.error(f"Ein Fehler ist beim Import Fees aus '{filename}' aufgetreten: {e}")
        return None    

def taxes_import(filename, logger):
    try:
        # Lese die ersten vier Spalten aus der Excel-Datei
        df = pd.read_excel(filename, usecols=[0, 1, 2, 6], names=['date', 'wkn', 'bank', 'delta', 'invest_divest', 'fees', 'taxes'])

        # Konvertiere 'wkn' und 'bank' in Kleinbuchstaben
        df['wkn'] = df['wkn'].str.lower()
        df['bank'] = df['bank'].str.lower()

        # Entferne Zeilen mit NaN in 'wkn', 'bank' oder 'taxes'
        df.dropna(subset=['wkn', 'bank', 'taxes'], inplace=True)

        # Setze den MultiIndex auf 'date', 'wkn', 'bank'
        df.set_index(['date', 'wkn', 'bank'], inplace=True)

        # Fasse Einträge mit demselben MultiIndex zusammen und summiere die Wert-Spalte
        # Damit werden mehrere Transaktion an einem Tag für eine WKN (bei der gleichen bank) zu einem Eintrag kombiniert
        df = df.groupby(level=['date', 'wkn', 'bank']).sum()

        return df

    except FileNotFoundError:
        logger.error(f"Die Datei '{filename}' wurde nicht gefunden.")
        return None
    except Exception as e:
        logger.error(f"Ein Fehler ist beim Import Taxes aus '{filename}' aufgetreten: {e}")
        return None    

def interest_dividends_import(filename, logger):
    try:
        # Lese die ersten vier Spalten aus der Excel-Datei
        df = pd.read_excel(filename, usecols=[0, 1, 2, 8], names=['date', 'wkn', 'bank', 'delta', 'invest_divest', 'fees', 'taxes', 'transaction_value_at_price', 'interest_dividends'])

        # Konvertiere 'wkn' und 'bank' in Kleinbuchstaben
        df['wkn'] = df['wkn'].str.lower()
        df['bank'] = df['bank'].str.lower()

        # Entferne Zeilen mit NaN in 'wkn', 'bank' oder 'interest_dividends'
        df.dropna(subset=['wkn', 'bank', 'interest_dividends'], inplace=True)

        # Setze den MultiIndex auf 'date', 'wkn', 'bank'
        df.set_index(['date', 'wkn', 'bank'], inplace=True)

        # Fasse Einträge mit demselben MultiIndex zusammen und summiere die Wert-Spalte
        # Damit werden mehrere Transaktion an einem Tag für eine WKN (bei der gleichen bank) zu einem Eintrag kombiniert
        df = df.groupby(level=['date', 'wkn', 'bank']).sum()

        return df

    except FileNotFoundError:
        logger.error(f"Die Datei '{filename}' wurde nicht gefunden.")
        return None
    except Exception as e:
        logger.error(f"Ein Fehler ist beim Import Interst and Dividends aus '{filename}' aufgetreten: {e}")
        return None    

def transaction_value_at_price_import(filename, logger):
    try:
        # Lese die ersten vier Spalten aus der Excel-Datei
        df = pd.read_excel(filename, usecols=[0, 1, 2, 7], names=['date', 'wkn', 'bank', 'delta', 'invest_divest', 'fees', 'taxes', 'transaction_value_at_price', 'interest_dividends'])

        # Konvertiere 'wkn' und 'bank' in Kleinbuchstaben
        df['wkn'] = df['wkn'].str.lower()
        df['bank'] = df['bank'].str.lower()

        # Entferne Zeilen mit NaN in 'wkn', 'bank' oder 'interest_dividends'
        df.dropna(subset=['wkn', 'bank', 'transaction_value_at_price'], inplace=True)

        # Setze den MultiIndex auf 'date', 'wkn', 'bank'
        df.set_index(['date', 'wkn', 'bank'], inplace=True)

        # Fasse Einträge mit demselben MultiIndex zusammen und summiere die Wert-Spalte
        # Damit werden mehrere Transaktion an einem Tag für eine WKN (bei der gleichen bank) zu einem Eintrag kombiniert
        df = df.groupby(level=['date', 'wkn', 'bank']).sum()

        return df

    except FileNotFoundError:
        logger.error(f"Die Datei '{filename}' wurde nicht gefunden.")
        return None
    except Exception as e:
        logger.error(f"Ein Fehler ist beim Import Interst and Dividends aus '{filename}' aufgetreten: {e}")
        return None    

# Funktion zum Prüfung ob alle WKN in Buchungsdaten in Instruments gelistet sind
def bookings_check_for_instruments(bookings, instruments):
    """
    Überprüft, ob alle wkns aus 'bookings' im DataFrame 'instruments' vorhanden sind und gibt eine Liste fehlender wkns zurück.
    
    Parameter:
        bookings (DataFrame): Der DataFrame mit Buchungsdaten, der eine wkn-Spalte oder -Index enthalten muss.
        instruments (DataFrame): Der DataFrame mit Instrumenten-Daten, der eine wkn-Spalte oder -Index enthalten muss.
        
    Rückgabe:
        missing_in_instruments (list): Liste der wkns aus 'bookings', die nicht in 'instruments' enthalten sind.
    """
    # Extrahiere die wkns aus dem bookings DataFrame
    wkn_bookings = set(bookings.index.get_level_values('wkn'))
    wkn_instruments = set(instruments.index)

    # Erstelle die Liste der wkns in bookings, die in instruments fehlen
    missing_in_instruments = list(wkn_bookings - wkn_instruments)
    
    return missing_in_instruments

# Funktion zur Umsetzung der Buchungen in ein Bestandsfile für alle Tage
def shares_from_bookings(bookings, end_date, logger):
    """
    Erweitert den DataFrame `bookings` mit allen Kombinationen von Datum, wkn und Bank
    bis zu einem angegebenen Enddatum und berechnet die laufende Summe.
    debei bedeutet share die Anzahl der Anteile für eine WKN an

    Parameter:
        bookings (DataFrame): Ein DataFrame mit MultiIndex (date, wkn, bank) und einer Spalte 'delta'.
        end_date (datetime): Das Datum, bis zu dem der DataFrame aufgebaut werden soll.

    Rückgabe:
        DataFrame: Ein erweiterter DataFrame mit dem MultiIndex (date, wkn, bank) und der laufenden Summe in der Spalte 'share'.
    """
    # Bestimme das vollständige Datumsspektrum bis zum übergebenen Enddatum
    all_dates = pd.date_range(bookings.index.get_level_values('date').min(), end_date)
    wkns = bookings.index.get_level_values('wkn').unique()
    banks = bookings.index.get_level_values('bank').unique()
    
    # Erstelle einen vollständigen MultiIndex für Datum, wkn und Bank
    full_index = pd.MultiIndex.from_product([all_dates, wkns, banks], names=['date', 'wkn', 'bank'])
    
    # Reindexiere den DataFrame, um alle Kombinationen von Datum, wkn und Bank abzudecken, und fülle NaN mit 0
    bookings_expanded = bookings.reindex(full_index).fillna(0)
    
    # Berechne die laufende Summe über das Datum für jede Kombination von wkn und Bank
    bookings_expanded['delta'] = bookings_expanded.groupby(['wkn', 'bank'])['delta'].cumsum()
    
    # Setze alle Werte kleiner als 0.0001 auf 0
    bookings_expanded['delta'] = bookings_expanded['delta'].where(bookings_expanded['delta'] >= 0.0001, 0)
    
    # Benenne die Spalte 'delta' in 'share' um
    bookings_expanded.rename(columns={'delta': 'share'}, inplace=True)

    logger.info('Positionen (shares) auf Tagesbasis erfolgreich aufgebaut')
    
    return bookings_expanded

# Funktion zur Berechnung der Wertbestände values (pro WKN) aus den Beständen shares (Stück) und Kursen prices
def values_from_shares_and_prices(shares_day_banks, prices):
    """
    Multipliziert die Positionen und Preise für jeden Indexwert (date, wkn, bank) und gibt das Ergebnis zurück.
    
    Parameter:
        shares_day_banks (DataFrame): DataFrame mit MultiIndex (date, wkn, bank) und einer 'share' Spalte.
        prices (DataFrame): DataFrame mit MultiIndex (date, wkn) und einer 'price' Spalte.
        
    Rückgabe:
        values (DataFrame): DataFrame mit MultiIndex (date, wkn, bank) und dem Ergebnis der Multiplikation 'value'.
    """
    # Erweitere den prices DataFrame um den Index 'bank'
    banks = shares_day_banks.index.get_level_values('bank').unique()
    prices_expanded = prices.reindex(pd.MultiIndex.from_product(
        [prices.index.get_level_values('date').unique(),
         prices.index.get_level_values('wkn').unique(),
         banks],
        names=['date', 'wkn', 'bank']
    ))

    # Multipliziere die Werte in 'share' und 'price' für jeden Indexwert
    values = shares_day_banks.copy()
    values['value'] = values['share'] * prices_expanded['price']

    return values[['value']]

# Transformiert 3D Datafram mit Multiindex (date, wkn, bank) in 2D Dataframe (date, wkn)
def aggregate_banks(df):
    """
    Aggregiert die Werte in einem DataFrame mit MultiIndex (date, wkn, bank) über alle Banken
    für jede Kombination von date und wkn.

    Parameter:
        df (DataFrame): Ein DataFrame mit MultiIndex (date, wkn, bank) und den aggregierbaren Werten.

    Rückgabe:
        DataFrame: Aggregierter DataFrame mit MultiIndex (date, wkn) und den aggregierten Werten.
    """
    # Prüfe, ob der DataFrame den erwarteten MultiIndex (date, wkn, bank) hat
    expected_index = ['date', 'wkn', 'bank']
    if list(df.index.names) != expected_index:
        raise ValueError(f"Der DataFrame muss den MultiIndex {expected_index} haben.")

    # Aggregiere die Werte für jede Kombination von date und wkn über alle Banken
    df_aggregated = df.groupby(['date', 'wkn']).sum()

    return df_aggregated

def gains_losses_before_fees_taxes_day(values_day_df, transaction_value_at_price_day_df):
    """
    Berechnet Tages-Gewinne/-Verluste vor Gebühren und Steuern:
    (Depotwert heute - Depotwert gestern) + Transaktionswert heute

    Parameter:
        values_day_df (pd.DataFrame): MultiIndex (day, wkn), Spalte: Depotwert
        transaction_value_at_price_day_df (pd.DataFrame): MultiIndex (day, wkn), Spalte: Transaktionswert

    Rückgabe:
        pd.DataFrame: MultiIndex (day, wkn), Spalte 'gains_losses_before_fee_taxes'
    """

    # Spaltennamen merken
    value_col = values_day_df.columns[0]
    trans_col = transaction_value_at_price_day_df.columns[0]

    # 1. Werte nach Datum und WKN sortieren (wichtig für shift)
    values_day_df = values_day_df.sort_index()
    transaction_value_at_price_day_df = transaction_value_at_price_day_df.sort_index()

    # 2. Depotwert gestern berechnen (je WKN)
    values_yesterday_df = values_day_df.groupby(level='wkn').shift(1)
    values_yesterday_df.columns = ['value_yesterday']

    # 3. Umbenennen der Originalspalten
    values_today_df = values_day_df.rename(columns={value_col: 'value_today'})
    transaction_df = transaction_value_at_price_day_df.rename(columns={trans_col: 'transaction_today'})

    # 4. Zusammenführen (Outer Join über alle Indizes)
    combined = values_today_df.join(values_yesterday_df, how='outer') \
                              .join(transaction_df, how='outer')

    # 5. Fehlende Werte mit 0 ersetzen
    combined = combined.fillna(0)

    # 6. Formel anwenden
    combined['gains_losses_before_fees_taxes'] = (
        combined['value_today'] - combined['value_yesterday'] + combined['transaction_today']
    )

    # 7. Nur das Ergebnis zurückgeben
    return combined[['gains_losses_before_fees_taxes']]




def unrealized_gains_losses_day(shares, prices):
    """
    Berechnet die tägliche Gewinn-Verlust-Matrix für jede Kombination von date und wkn
    es werden nur Gewinne und Verluste aus Kurs-Veränderungen berücksichtig, keine Gebühren, Steuern oder Zinsen/Divdidend
    anhand von 'share' und 'price'. NaN-Werte werden durch 0 ersetzt.

    Parameter:
        shares (DataFrame): DataFrame mit MultiIndex (date, wkn) und einer Spalte 'share'.
        prices (DataFrame): DataFrame mit MultiIndex (date, wkn) und einer Spalte 'price'.
        
    Rückgabe:
        DataFrame: DataFrame mit MultiIndex (date, wkn) und der täglichen Gewinn-Verlust-Werte in der Spalte 'unrealized_gains_losses'.
    """
    # Berechne die tägliche Preisänderung: price(date) - price(date - 1) für jede wkn
    price_diff = prices.groupby('wkn')['price'].diff().fillna(0)

    # Berechne die Gewinn-Verlust-Werte: share * price_diff für jeden Indexwert (date, wkn)
    unrealized_gains_losses = (shares['share'].fillna(0) * price_diff).fillna(0)

    # Erstelle den Ergebnis-DataFrame
    unrealized_gains_losses_df = unrealized_gains_losses.to_frame(name='unrealized_gains_losses')
    
    return unrealized_gains_losses_df

def realized_gains_losses_day(fees_df, taxes_df, interests_dividends_df):
    """
    Führt drei DataFrames mit identischem MultiIndex (date, wkn) zusammen
    und berechnet die Summe je (date, wkn) in der Spalte 'realized_gains_losses'.

    Parameter:
        fees_df (pd.DataFrame): DataFrame mit MultiIndex (date, wkn) und einer Werte-Spalte.
        taxes_df (pd.DataFrame): Gleich aufgebaut.
        interests_dividends_df (pd.DataFrame): Gleich aufgebaut.

    Rückgabe:
        pd.DataFrame: Ergebnis mit Spalte 'realized_gains_losses'.
    """

    # Umbenennen der Spalten zur Unterscheidung (falls nicht schon geschehen)
    fees_df = fees_df.rename(columns={fees_df.columns[0]: 'fees'})
    taxes_df = taxes_df.rename(columns={taxes_df.columns[0]: 'taxes'})
    interests_dividends_df = interests_dividends_df.rename(columns={interests_dividends_df.columns[0]: 'interests_dividends'})

    # Join aller drei DataFrames auf gemeinsamen MultiIndex (outer join)
    combined_df = fees_df.join(taxes_df, how='outer').join(interests_dividends_df, how='outer')

    # Fehlende Werte durch 0 ersetzen
    combined_df = combined_df.fillna(0)

    # Neue Spalte mit der Summe
    combined_df['realized_gains_losses'] = combined_df['fees'] + combined_df['taxes'] + combined_df['interests_dividends']

    # Nur die Ergebnis-Spalte zurückgeben
    return combined_df[['realized_gains_losses']]

def invest_day(filename, start_date, end_date, logger):
    """
    Liest die Buchungsdaten aus der angegebenen Datei ein und verarbeitet sie, um Investitionsdaten (Einschuss/Entnahme)
    für jeden Tag innerhalb des Zeitraums von start_date bis end_date zu erhalten.

    Parameter:
        filename (str): Der Dateipfad zur Excel-Datei.
        start_date (str oder datetime): Das Startdatum für den erweiterten Datumsbereich.
        end_date (str oder datetime): Das Enddatum für den erweiterten Datumsbereich.
        logger (ExtendedLogger): Logger instance for output and logging.

    Rückgabe:
        DataFrame: Ein DataFrame mit dem Index 'date' und einer Spalte 'invest', die die täglichen Investitionswerte enthält.
    """
    try:
        # Lese die Spalten 'date', 'delta' und 'invest_divest' aus der Datei ein
        df = pd.read_excel(filename, usecols=['date', 'delta', 'invest_divest'])

        # Setze die Spalte 'date' als Index
        df.set_index('date', inplace=True)

        # Entferne alle Zeilen ohne Eintrag in der Spalte 'invest_divest'
        df = df.dropna(subset=['invest_divest'])

        # Lösche die Spalte 'invest_divest'
        df.drop(columns='invest_divest', inplace=True)

        # Aggregiere doppelte Datumswerte, indem die 'delta'-Werte summiert werden
        df = df.groupby('date').sum()

        # Erweitere den DataFrame auf den gesamten Datumsbereich von start_date bis end_date
        all_dates = pd.date_range(start=start_date, end=end_date)
        df = df.reindex(all_dates, fill_value=0)

        # Benenne die Spalte 'delta' in 'invest' um
        df.rename(columns={'delta': 'invest'}, inplace=True)

        return df

    except FileNotFoundError:
        logger.error(f"Die Datei '{filename}' wurde nicht gefunden (invest_day).")
        return None
    except Exception as e:
        logger.error(f"Ein Fehler ist aufgetreten: {e} (invest_day)")
        return None

def values_type_month(values_month, instruments_type):
    """
    Berechnet den monatlichen Wert pro Typ auf Basis der Daten in values_month und instruments_type.
    Entfernt die laufende Nummer, lässt nur den spezifischen Typ (z.B. "aktie" statt "type aktie") in der Spalte 'type' stehen.

    Parameter:
        values_month (DataFrame): DataFrame mit MultiIndex (date, wkn) und Spalte 'value'.
        instruments_type (DataFrame): DataFrame mit MultiIndex (wkn, type) und Spalte 'share'.

    Rückgabe:
        DataFrame: Ein DataFrame mit MultiIndex (date, type) und Spalte 'type_value'.
    """
    
    # Bereinige den 'type'-Index in instruments_type, um Präfix 'type' zu entfernen
    instruments_type.index = instruments_type.index.set_levels(
        instruments_type.index.levels[1].str.replace('type ', '', regex=False),
        level=1
    )

    # Führe das Merge der beiden DataFrames basierend auf 'wkn' durch
    merged_df = values_month.reset_index().merge(
        instruments_type.reset_index(), 
        on='wkn', 
        how='left'
    )

    # Berechne type_value als das Produkt von 'value' und 'share'
    merged_df['type_value'] = merged_df['value'] * merged_df['share']

    # Gruppiere nach 'date' und 'type', und summiere 'type_value'
    result_df = merged_df.groupby(['date', 'type'])['type_value'].sum().reset_index()

    # Setze den MultiIndex (date, type) und entferne die laufende Nummer
    result_df.set_index(['date', 'type'], inplace=True)

    return result_df

def values_region_month(values_month, instruments_region, logger):
    """
    Berechnet den monatlichen Wert pro Region auf Basis der Daten in values_month und instruments_region.
    Entfernt die laufende Nummer, lässt nur die spezifische Region in der Spalte 'region' stehen.

    Parameter:
        values_month (DataFrame): DataFrame mit MultiIndex (date, wkn) und Spalte 'value'.
        instruments_region (DataFrame): DataFrame mit MultiIndex (wkn, region) und Spalte 'share'.
        logger (ExtendedLogger): Logger instance for output and logging.

    Rückgabe:
        DataFrame: Ein DataFrame mit MultiIndex (date, region) und Spalte 'region_value'.
    """
    
    # Bereinige den 'region'-Index in instruments_region, um Präfix 'Reg' zu entfernen
    instruments_region.index = instruments_region.index.set_levels(
        instruments_region.index.levels[1].str.replace('reg ', '', regex=False),
        level=1
    )

    # Führe das Merge der beiden DataFrames basierend auf 'wkn' durch
    merged_df = values_month.reset_index().merge(
        instruments_region.reset_index(), 
        on='wkn', 
        how='left'
    )
    
    # Berechne region_value als das Produkt von 'value' und 'share'
    merged_df['region_value'] = merged_df['value'] * merged_df['share']

    # Gruppiere nach 'date' und 'region', und summiere 'region_value'
    if 'reg' not in merged_df.columns:
        logger.error("Keine Spalten mit 'reg' gefunden. Bitte prüfen Sie die Eingabedaten.")
        return None
    
    result_df = merged_df.groupby(['date', 'reg'])['region_value'].sum().reset_index()

    # Setze den MultiIndex (date, region) und entferne die laufende Nummer
    result_df.set_index(['date', 'reg'], inplace=True)

    return result_df

# Tagespofitabilität aus Wertänderung der Tageswert
def yield_day_from_values_day(gains_losses_before_fees_taxes_day_df, values_day_df, logger):

    """
    Berechnet die tägliche Rendite (Yield) pro WKN und Tag.

    Parameter:
        unrealized_gains_losses_day (DataFrame): DataFrame mit MultiIndex (date, wkn) und Spalte 'unrealized_gains_losses'.
        values_day (DataFrame): DataFrame mit MultiIndex (date, wkn) und Spalte 'value'.
        logger (ExtendedLogger): Logger instance for output and logging.

    Rückgabe:
        DataFrame: DataFrame mit MultiIndex (date, wkn) und Spalte 'yield', die die tägliche Rendite angibt.
    """
    try:
        # Nur Werte verwenden, die im Index von values_day enthalten sind
        gains_loss_filtered_df = gains_losses_before_fees_taxes_day_df.loc[
            gains_losses_before_fees_taxes_day_df.index.intersection(values_day_df.index)
        ]

        # Werte-DataFrame kopieren (für Sicherheit)
        values_filtered_df = values_day_df.copy()

        # Berechnung: Gewinn/Verlust relativ zum Depotwert
        yield_series = (
            gains_loss_filtered_df['gains_losses_before_fees_taxes'] / values_filtered_df['value']
        ).replace([np.inf, -np.inf], np.nan).fillna(0)

        return yield_series.to_frame(name='yield')
    
    except Exception as e:
        logger.error(f"Fehler beim Berechnen der täglichen Profitabilität: {e}")
        return None


# Kumulierte Jahres Profitabilität über TWR Formel: (Produkt Reihe (1+Tagesprofitabilität))-1
def yield_year_from_values_day(yield_excl_div_day, values_day, logger):
    """
    Berechnet die kumulierte Rendite vom ersten bis zum letzten Tag eines Jahres je WKN.
    Gruppiert wird nach dem jeweils letzten berücksichtigten Datum je Jahr und WKN.

    Parameter:
        yield_excl_div_day (DataFrame): MultiIndex (date, wkn), Spalte 'yield'
        values_day (DataFrame): MultiIndex (date, wkn), Spalte 'value'
        logger (ExtendedLogger): Logger instance for output and logging.

    Rückgabe:
        DataFrame: MultiIndex (last_date_per_year, wkn), Spalte 'annual_yield'
    """
    try:
        # Index sicherstellen: datetime + wkn
        yield_excl_div_day = yield_excl_div_day.copy()
        yield_excl_div_day.index = pd.MultiIndex.from_arrays([
            pd.to_datetime(yield_excl_div_day.index.get_level_values('date')),
            yield_excl_div_day.index.get_level_values('wkn')
        ], names=['date', 'wkn'])

        # Join mit values_day (für Bestandsprüfung)
        yield_and_value = yield_excl_div_day.join(values_day[['value']], how='inner')

        # Nur Tage mit positivem Bestand
        yield_and_value = yield_and_value[yield_and_value['value'] > 0]

        # Extrahiere Jahr separat für Gruppierung
        yield_and_value['year'] = yield_and_value.index.get_level_values('date').year

        # Ermittle pro (Jahr, WKN) das letzte Datum
        last_dates = (
            yield_and_value.reset_index()
            .groupby(['year', 'wkn'])['date']
            .max()
            .reset_index()
            .rename(columns={'date': 'last_date'})
        )

    

        # Merge, um jedem Tageswert den zugehörigen "Jahresendstempel" zu geben
        merged = yield_and_value.reset_index().merge(last_dates, on=['year', 'wkn'])

        # Gruppiere nach (last_date, wkn) statt (year, wkn)
        result = merged.groupby(['last_date', 'wkn'])['yield'].apply(
            lambda x: np.prod(1 + x) - 1
        )

        return result.to_frame(name='annual_yield')

    except Exception as e:
        logger.error(f"Fehler beim Berechnen der jährlichen kumulierten Rendite: {e}")
        return None


# Renditekomponenten detailliert: Kursrendite, Dividenden, Gebühren, Steuern, Gesamtrendite
def yield_components_day(gains_losses_before_fees_taxes_df, fees_df, taxes_df, interest_dividends_df, values_day_df, transaction_value_at_price_df, logger):
    """
    Berechnet die detaillierten Renditekomponenten pro WKN und Tag.
    Trennt die Gesamtrendite in ihre Bestandteile auf:
    - yield_price: Reine Kursrendite (ohne Dividenden, Gebühren, Steuern)
    - yield_dividends: Rendite aus Zinsen/Dividenden
    - yield_fees: Belastung durch Gebühren (negativ)
    - yield_taxes: Belastung durch Steuern (negativ)
    - yield_total: Gesamtrendite (Summe aller Komponenten)

    Bei großen Transaktionen (Käufe/Verkäufe) werden Gebühren und Steuern
    auf den Transaktionswert bezogen, nicht auf den End-Portfoliowert,
    um extreme Prozentsätze zu vermeiden.

    Parameter:
        gains_losses_before_fees_taxes_df (DataFrame): MultiIndex (date, wkn), Spalte 'gains_losses_before_fees_taxes'
        fees_df (DataFrame): MultiIndex (date, wkn), Spalte 'fees'
        taxes_df (DataFrame): MultiIndex (date, wkn), Spalte 'taxes'
        interest_dividends_df (DataFrame): MultiIndex (date, wkn), Spalte 'interest_dividends'
        values_day_df (DataFrame): MultiIndex (date, wkn), Spalte 'value'
        transaction_value_at_price_df (DataFrame): MultiIndex (date, wkn), Spalte 'transaction_value_at_price'
        logger (ExtendedLogger): Logger instance for output and logging.

    Rückgabe:
        DataFrame: MultiIndex (date, wkn) mit Spalten:
                   ['yield_price', 'yield_dividends', 'yield_fees', 'yield_taxes', 'yield_total']
    """
    try:
        # Validierung der Eingabedaten
        if values_day_df is None or values_day_df.empty:
            raise ValueError("values_day_df ist None oder leer - kann Rendite nicht berechnen")

        if gains_losses_before_fees_taxes_df is None or gains_losses_before_fees_taxes_df.empty:
            raise ValueError("gains_losses_before_fees_taxes_df ist None oder leer - kann Kursrendite nicht berechnen")

        if 'value' not in values_day_df.columns:
            raise ValueError(f"Spalte 'value' fehlt in values_day_df. Vorhandene Spalten: {values_day_df.columns.tolist()}")

        if 'gains_losses_before_fees_taxes' not in gains_losses_before_fees_taxes_df.columns:
            raise ValueError(f"Spalte 'gains_losses_before_fees_taxes' fehlt in gains_losses_before_fees_taxes_df. Vorhandene Spalten: {gains_losses_before_fees_taxes_df.columns.tolist()}")

        logger.info("Starte Berechnung der Renditekomponenten (yield_components_day)...")

        # Exclude 'cash' from profitability calculations
        values_day_filtered = values_day_df[values_day_df.index.get_level_values('wkn').str.lower() != 'cash']
        gains_losses_filtered = gains_losses_before_fees_taxes_df[gains_losses_before_fees_taxes_df.index.get_level_values('wkn').str.lower() != 'cash']

        # Filter other DataFrames if they exist
        interest_dividends_filtered = None
        if interest_dividends_df is not None and not interest_dividends_df.empty:
            interest_dividends_filtered = interest_dividends_df[interest_dividends_df.index.get_level_values('wkn').str.lower() != 'cash']

        fees_filtered = None
        if fees_df is not None and not fees_df.empty:
            fees_filtered = fees_df[fees_df.index.get_level_values('wkn').str.lower() != 'cash']

        taxes_filtered = None
        if taxes_df is not None and not taxes_df.empty:
            taxes_filtered = taxes_df[taxes_df.index.get_level_values('wkn').str.lower() != 'cash']

        transaction_value_filtered = None
        if transaction_value_at_price_df is not None and not transaction_value_at_price_df.empty:
            transaction_value_filtered = transaction_value_at_price_df[transaction_value_at_price_df.index.get_level_values('wkn').str.lower() != 'cash']

        logger.info("Cash ausgeschlossen von yield_components_day Berechnung")

        # Basis: values_day_df als Ausgangspunkt
        result = values_day_filtered[['value']].copy()

        # Join aller Komponenten
        result = result.join(gains_losses_filtered[['gains_losses_before_fees_taxes']], how='left')
        result = result.join(interest_dividends_filtered, how='left') if interest_dividends_filtered is not None else result
        result = result.join(fees_filtered, how='left') if fees_filtered is not None else result
        result = result.join(taxes_filtered, how='left') if taxes_filtered is not None else result
        result = result.join(transaction_value_filtered, how='left') if transaction_value_filtered is not None else result

        # Fehlende Werte mit 0 auffüllen (nur für die Komponenten, nicht für value)
        for col in ['gains_losses_before_fees_taxes', 'interest_dividends', 'fees', 'taxes', 'transaction_value_at_price']:
            if col not in result.columns:
                result[col] = 0.0
            else:
                result[col] = result[col].fillna(0.0)

        # Prüfen auf Division durch Null
        zero_values = result['value'] == 0
        if zero_values.any():
            count_zeros = zero_values.sum()
            logger.warning(f"{count_zeros} Einträge mit value=0 gefunden - setze Rendite auf 0.0")

        # Renditekomponenten berechnen (nur wo value > 0)
        result['yield_price'] = 0.0
        result['yield_dividends'] = 0.0
        result['yield_fees'] = 0.0
        result['yield_taxes'] = 0.0

        non_zero_mask = result['value'] > 0

        # Intelligente Nenner-Wahl für alle Rendite-Komponenten
        # Bei großen Transaktionen (|transaction| > 50% des Portfoliowerts) nutze Transaktionswert als Nenner
        # um extreme Prozentsätze bei fast vollständigen Verkäufen zu vermeiden
        abs_transaction = result['transaction_value_at_price'].abs()

        # Bestimme den Nenner: transaction_value wenn groß, sonst value
        denominator = result['value'].copy()
        large_transaction_mask = (abs_transaction > result['value'] * 0.5) & (abs_transaction > 0)
        denominator.loc[large_transaction_mask] = abs_transaction.loc[large_transaction_mask]

        # Berechne alle Komponenten mit intelligentem Nenner
        valid_denominator_mask = (denominator > 0) & non_zero_mask

        result.loc[valid_denominator_mask, 'yield_price'] = (
            result.loc[valid_denominator_mask, 'gains_losses_before_fees_taxes'] / denominator.loc[valid_denominator_mask]
        )
        result.loc[valid_denominator_mask, 'yield_dividends'] = (
            result.loc[valid_denominator_mask, 'interest_dividends'] / denominator.loc[valid_denominator_mask]
        )
        result.loc[valid_denominator_mask, 'yield_fees'] = (
            result.loc[valid_denominator_mask, 'fees'] / denominator.loc[valid_denominator_mask]
        )
        result.loc[valid_denominator_mask, 'yield_taxes'] = (
            result.loc[valid_denominator_mask, 'taxes'] / denominator.loc[valid_denominator_mask]
        )

        # Logging für große Transaktionen
        if large_transaction_mask.sum() > 0:
            logger.info(f"{large_transaction_mask.sum()} Einträge mit großen Transaktionen erkannt - nutze Transaktionswert als Nenner für alle Komponenten")

        # Gesamtrendite = Summe aller Komponenten
        result['yield_total'] = (
            result['yield_price'] +
            result['yield_dividends'] +
            result['yield_fees'] +      # fees sind bereits negativ
            result['yield_taxes']       # taxes sind bereits negativ
        )

        # Prüfen auf inf/-inf Werte
        inf_mask = np.isinf(result[['yield_price', 'yield_dividends', 'yield_fees', 'yield_taxes', 'yield_total']]).any(axis=1)
        if inf_mask.any():
            count_inf = inf_mask.sum()
            logger.error(f"{count_inf} Einträge mit inf/-inf Werten gefunden - dies deutet auf einen Fehler hin")
            raise ValueError(f"{count_inf} Einträge mit unendlichen Werten in Renditeberechnung - bitte Daten prüfen")

        # Nur Rendite-Spalten zurückgeben
        result_df = result[['yield_price', 'yield_dividends', 'yield_fees', 'yield_taxes', 'yield_total']]

        logger.info(f"Renditekomponenten erfolgreich berechnet: {len(result_df)} Einträge")
        logger.info(f"Durchschnittliche Gesamtrendite: {result_df['yield_total'].mean():.6f}")

        return result_df

    except ValueError as ve:
        logger.error(f"Validierungsfehler bei yield_components_day: {ve}")
        raise
    except KeyError as ke:
        logger.error(f"Spalten-Fehler bei yield_components_day: {ke}")
        raise
    except Exception as e:
        logger.error(f"Unerwarteter Fehler bei yield_components_day: {type(e).__name__}: {e}")
        raise


# Kumulierte Jahresrendite für alle Renditekomponenten
def yield_components_year(yield_components_day_df, values_day_df, logger):
    """
    Berechnet die kumulierte jährliche Rendite für alle Renditekomponenten.
    Verwendet die Time-Weighted Return (TWR) Formel: Produkt(1 + tägliche_rendite) - 1
    Gruppiert nach dem letzten Datum je Jahr und WKN.

    Parameter:
        yield_components_day_df (DataFrame): MultiIndex (date, wkn), Spalten:
                                             ['yield_price', 'yield_dividends', 'yield_fees', 'yield_taxes', 'yield_total']
        values_day_df (DataFrame): MultiIndex (date, wkn), Spalte 'value'
        logger (ExtendedLogger): Logger instance for output and logging.

    Rückgabe:
        DataFrame: MultiIndex (last_date_per_year, wkn), Spalten:
                   ['annual_yield_price', 'annual_yield_dividends', 'annual_yield_fees',
                    'annual_yield_taxes', 'annual_yield_total']
    """
    try:
        # Validierung der Eingabedaten
        if yield_components_day_df is None or yield_components_day_df.empty:
            raise ValueError("yield_components_day_df ist None oder leer")

        if values_day_df is None or values_day_df.empty:
            raise ValueError("values_day_df ist None oder leer")

        required_columns = ['yield_price', 'yield_dividends', 'yield_fees', 'yield_taxes', 'yield_total']
        missing_columns = [col for col in required_columns if col not in yield_components_day_df.columns]
        if missing_columns:
            raise ValueError(f"Fehlende Spalten in yield_components_day_df: {missing_columns}")

        logger.info("Starte Berechnung der jährlichen Renditekomponenten (yield_components_year)...")

        # Index sicherstellen: datetime + wkn
        yield_comp = yield_components_day_df.copy()
        yield_comp.index = pd.MultiIndex.from_arrays([
            pd.to_datetime(yield_comp.index.get_level_values('date')),
            yield_comp.index.get_level_values('wkn')
        ], names=['date', 'wkn'])

        # Join mit values_day (für Bestandsprüfung)
        yield_and_value = yield_comp.join(values_day_df[['value']], how='inner')

        # Nur Tage mit positivem Bestand
        yield_and_value = yield_and_value[yield_and_value['value'] > 0]

        if yield_and_value.empty:
            raise ValueError("Keine Einträge mit positivem Bestand gefunden")

        # Extrahiere Jahr separat für Gruppierung
        yield_and_value['year'] = yield_and_value.index.get_level_values('date').year

        # Ermittle pro (Jahr, WKN) das letzte Datum
        last_dates = (
            yield_and_value.reset_index()
            .groupby(['year', 'wkn'])['date']
            .max()
            .reset_index()
            .rename(columns={'date': 'last_date'})
        )

        # Merge, um jedem Tageswert den zugehörigen "Jahresendstempel" zu geben
        merged = yield_and_value.reset_index().merge(last_dates, on=['year', 'wkn'])

        # Gruppiere nach (last_date, wkn) und berechne TWR für jede Komponente
        def calculate_twr(series):
            """Time-Weighted Return: Produkt(1 + r) - 1"""
            return np.prod(1 + series) - 1

        grouped = merged.groupby(['last_date', 'wkn'])

        result = pd.DataFrame({
            'annual_yield_price': grouped['yield_price'].apply(calculate_twr),
            'annual_yield_dividends': grouped['yield_dividends'].apply(calculate_twr),
            'annual_yield_fees': grouped['yield_fees'].apply(calculate_twr),
            'annual_yield_taxes': grouped['yield_taxes'].apply(calculate_twr),
            'annual_yield_total': grouped['yield_total'].apply(calculate_twr)
        })

        logger.info(f"Jährliche Renditekomponenten erfolgreich berechnet: {len(result)} Jahres-WKN-Kombinationen")
        logger.info(f"Durchschnittliche jährliche Gesamtrendite: {result['annual_yield_total'].mean():.6f}")

        return result

    except ValueError as ve:
        logger.error(f"Validierungsfehler bei yield_components_year: {ve}")
        raise
    except Exception as e:
        logger.error(f"Unerwarteter Fehler bei yield_components_year: {type(e).__name__}: {e}")
        raise


def profitability_year_table(yield_components_day_df, values_day_df, logger):
    """
    Erstellt eine Profitabilitäts-Tabelle mit Jahren als Zeilen und verschachtelten Spalten pro WKN.
    Für jede WKN gibt es zwei Spalten: {wkn}_days (Anzahl gehaltene Tage) und {wkn}_yield (Jahresrendite).

    Parameter:
        yield_components_day_df (DataFrame): MultiIndex (date, wkn), enthält 'yield_total'
        values_day_df (DataFrame): MultiIndex (date, wkn), enthält 'value'
        logger (ExtendedLogger): Logger instance for output and logging.

    Rückgabe:
        DataFrame: Index = Jahr (int), Spalten = {wkn}_days, {wkn}_yield für jede WKN
    """
    try:
        logger.info("Starte Erstellung der Jahres-Profitabilitäts-Tabelle...")

        # Validierung
        if yield_components_day_df is None or yield_components_day_df.empty:
            raise ValueError("yield_components_day_df ist None oder leer")
        if values_day_df is None or values_day_df.empty:
            raise ValueError("values_day_df ist None oder leer")

        # Kopie erstellen und Index sicherstellen
        yield_comp = yield_components_day_df.copy()
        yield_comp.index = pd.MultiIndex.from_arrays([
            pd.to_datetime(yield_comp.index.get_level_values('date')),
            yield_comp.index.get_level_values('wkn')
        ], names=['date', 'wkn'])

        # Exclude 'cash' from profitability calculations
        yield_comp = yield_comp[yield_comp.index.get_level_values('wkn').str.lower() != 'cash']
        logger.info("Cash ausgeschlossen von Profitabilitäts-Berechnung")

        # Join mit values (nur Tage mit Bestand)
        yield_and_value = yield_comp.join(values_day_df[['value']], how='inner')
        yield_and_value = yield_and_value[yield_and_value['value'] > 0]

        if yield_and_value.empty:
            raise ValueError("Keine Einträge mit positivem Bestand gefunden")

        # Extrahiere Jahr
        yield_and_value['year'] = yield_and_value.index.get_level_values('date').year

        # Berechne Tage pro Jahr-WKN
        df_reset = yield_and_value.reset_index()

        days_per_year_wkn = (
            df_reset.groupby(['year', 'wkn'])
            .size()
            .reset_index(name='days_held')
        )

        # Berechne TWR pro Jahr-WKN
        def calculate_twr(series):
            """Time-Weighted Return: Produkt(1 + r) - 1"""
            return np.prod(1 + series) - 1

        yield_per_year_wkn = (
            df_reset.groupby(['year', 'wkn'])['yield_total']
            .apply(calculate_twr)
            .reset_index(name='annual_yield')
        )

        # Merge days und yield
        combined = days_per_year_wkn.merge(yield_per_year_wkn, on=['year', 'wkn'])

        # Pivot: Jahr als Index, WKN als Spalten (mit days und yield)
        result_list = []

        for wkn in sorted(combined['wkn'].unique()):
            wkn_data = combined[combined['wkn'] == wkn][['year', 'days_held', 'annual_yield']]
            wkn_data = wkn_data.rename(columns={
                'days_held': f'{wkn}_days',
                'annual_yield': f'{wkn}_yield'
            })
            wkn_data = wkn_data.set_index('year')
            result_list.append(wkn_data)

        # Alle WKN-Dataframes zusammenführen
        result = pd.concat(result_list, axis=1)

        # Sortiere Spalten: Für jede WKN erst _days, dann _yield
        wkns_sorted = sorted(combined['wkn'].unique())
        sorted_columns = []
        for wkn in wkns_sorted:
            sorted_columns.append(f'{wkn}_days')
            sorted_columns.append(f'{wkn}_yield')

        result = result[sorted_columns]

        # Index-Name setzen
        result.index.name = 'year'

        logger.info(f"Jahres-Profitabilitäts-Tabelle erstellt: {len(result)} Jahre, {len(wkns_sorted)} WKNs")
        logger.info(f"Jahre: {sorted(result.index.tolist())}")

        return result

    except ValueError as ve:
        logger.error(f"Validierungsfehler bei profitability_year_table: {ve}")
        raise
    except Exception as e:
        logger.error(f"Unerwarteter Fehler bei profitability_year_table: {type(e).__name__}: {e}")
        raise


# Main Block 01: Initializing    
def initializing(settings_file, screen):
    """
    Initialisiert das Programm, indem das Arbeitsverzeichnis gesetzt wird, die Einstellungen geladen werden
    und die Verfügbarkeit der erforderlichen Dateien überprüft wird.

    Parameter:
        screen (bool): Wenn True, werden Bildschirmmeldungen angezeigt.
        log (bool): Wenn True, werden Meldungen ins Logfile geschrieben.

    Rückgabe:
        settings (dict): Ein Dictionary mit den Programmeinstellungen oder None bei Fehler.
    """
    error_count = 0
    warning_count = 0
    settings = None
    screen = True # Debug   

    # 1. Arbeitsverzeichnis setzen, kann auch einen benutzerdefinierten Pfad akzeptieren
    try:
        # Create temporary logger for early directory operations
        temp_logger = create_extended_logger("depot_init.log", screen, script_name='depot_init')
        set_working_directory("default", temp_logger)
        if screen:
            print("Info: Arbeitsverzeichnis initial auf Ausführungsordner gesetzt.")
    except Exception as e:
        if screen:
            print(f"ERROR: Fehler beim Setzen des Arbeitsverzeichnisses: {e}")
        error_count += 1
        # Fehlerergebnis melden und beenden
        if screen:
            print("ERROR: Initialisierung fehlgeschlagen. Das Programm wird beendet.")
        return None

    # 2. Einstellungen aus der Datei 'depot_file_settings.txt' lesen
    settings = settings_import(settings_file, temp_logger)
    if settings is None:
        if screen:
            print("ERROR: Einstellungen konnten nicht geladen werden.")
        error_count += 1
        # Fehlerergebnis melden und beenden
        if screen:
            print("ERROR: Initialisierung fehlgeschlagen. Das Programm wird beendet.")
        return None

    # 3. Arbeitsverzeichnis auf Einstellung aus Settings setzen
    try:
        # Use the same temporary logger for directory operations
        path = (settings or {}).get('Paths', {}).get('path', '')
        if not path:  # If path is empty or None, use default
            path = "default"
        set_working_directory(path, temp_logger)
        if screen:
            print("Info: Arbeitsverzeichnis erfolgreich gesetzt.")
    except Exception as e:
        if screen:
            print(f"ERROR: Fehler beim Setzen des Arbeitsverzeichnisses: {e}")
        error_count += 1
        # Fehlerergebnis melden und beenden
        if screen:
            print("ERROR: Initialisierung fehlgeschlagen. Das Programm wird beendet.")
        return None

    # 4.1. Logfile-Pfad aus den Einstellungen extrahieren
    logfile = (settings or {}).get('Files', {}).get('logfile')
    
    # 4.2. Prüfen, ob logfile None ist, und ggf. auf Standard setzen
    if logfile is None:
        logfile = 'logfile.txt'
        if screen:
            print("ERROR: Kein Logfile angegeben. Fallback auf 'logfile.txt'.")

    # 4.3. Prüfen, ob logfile existiert; wenn nicht, die Datei mit UTF-8 anlegen
    if not os.path.exists(logfile):
        try:
            with open(logfile, 'w', encoding='utf-8') as log_file:
                log_file.write("")  # Leere Datei anlegen
            if screen:
                print(f"Info: Logfile '{logfile}' wurde neu angelegt.")
        except Exception as e:
            if screen:
                print(f"ERROR: Logfile '{logfile}' konnte nicht erstellt werden: {e}")
            error_count += 1



    # 5. Überprüfen, ob die erforderlichen Dateien verfügbar sind
    file_list = list(settings['Files'].values())
    if not files_availability_check(file_list, temp_logger):
        if screen:
            print("ERROR: Eine oder mehrere Dateien fehlen.")
        error_count += 1
    else:
        if screen:
            print("Info: Alle Dateien verfügbar und erfolgreich geladen.")

    # Ergebnis ausgeben
    if error_count > 0:
        if screen:
            print("ERROR: Initialisierung fehlgeschlagen. Das Programm wird beendet.")
        sys.exit(1)
    elif warning_count > 0:
        if screen:
            print(f"WARNING: Initialisierung abgeschlossen mit {warning_count} Warnung(en).")
    else:
        if screen:
            print("Info: Initialisierung erfolgreich abgeschlossen.")
    return settings

# Main Block 02: Instrumente importieren
def instruments_import_and_process(settings, logger):
    """
    Importiert und verarbeitet die Instrumentendaten, indem die Haupt-Instruments-Datei, die Regions-Daten und die
    Typen-Daten eingelesen werden. Überprüft, ob die Dateien korrekt geladen wurden.

    Parameter:
        settings (dict): Die Programmeinstellungen.
        logfile (str): Der Name des Logfiles.
        screen (bool): Wenn True, werden Bildschirmmeldungen angezeigt.
        log (bool): Wenn True, werden Meldungen ins Logfile geschrieben.
        
    Rückgabe:
        tuple: Ein Tuple mit den DataFrames (instruments_df, instruments_region_df, instruments_type_df) oder
               (None, None, None) bei Fehlern.
    """
    error_count = 0
    warning_count = 0
    instruments_df = None
    instruments_region_df = None
    instruments_type_df = None

    try:
        # 2.1 Instruments-Datei importieren
        instruments_file = (settings or {}).get('Files', {}).get('instruments', '')
        instruments_df = instruments_import(instruments_file, logger)

        if instruments_df is None:
            logger.error(f"Fehler beim Import der Instruments-Datei '{instruments_file}'.")
            error_count += 1
            return None, None, None
        
        logger.info("Instruments-Datei erfolgreich importiert.")

        # 2.2 Instruments-Region-Daten importieren
        instruments_region_df = instruments_details_import(instruments_file, search_param="Reg", logger=logger)

        if instruments_region_df is None:
            logger.error("Eine oder mehrere WKN-Zeilen in 'instruments_region_df' ergeben nicht 100% .")
            error_count += 1
            return None, None, None

        logger.info("Instruments-Region-Daten erfolgreich importiert.")

        # 2.3 Instruments-Type-Daten importieren
        instruments_type_df = instruments_details_import(instruments_file, search_param="Type", logger=logger)

        if instruments_type_df is None:
            logger.error("Eine oder mehrere WKN-Zeilen in 'instruments_type_df' ergeben nicht 100%.")
            error_count += 1
            return None, None, None

        logger.info("Instruments-Type-Daten erfolgreich importiert.")

        # Exportiere instruments_type_df als Excel-Datei
        if (settings or {}).get('Export', {}).get('instruments_type_to_excel', {}):
            try:
                if (settings or {}).get("Export", {}).get("instruments_type_to_excel", {}).get("enabled", False):
                    export_df_to_excel(instruments_type_df, (settings or {}).get("Export", {}).get("instruments_type_to_excel", {}).get("filename", ""), logger)
            except Exception as e:
                logger.warning(f"Fehler beim Exportieren der Instruments-Type-Daten: {e}")
                warning_count += 1

    except Exception as e:
        logger.error(f"Unerwarteter Fehler beim Import der Typen der Instrumente: {e}")
        error_count += 1

    # Ergebnis ausgeben
    if error_count > 0:
        logger.error("Instruments-Import und -Verarbeitung fehlgeschlagen.")
    elif warning_count > 0:
        logger.warning(f"Instruments-Import und -Verarbeitung abgeschlossen mit {warning_count} Warnung(en).")
    else:
        logger.info("Instruments-Import und -Verarbeitung erfolgreich abgeschlossen.")
    return instruments_df, instruments_region_df, instruments_type_df

# Main Block 03: Kurse (prices) importieren, prüfen und updaten
def prices_import_and_process(settings, instruments_df, logger):
    """
    Importiert und verarbeitet die Preise, indem die Price-Datei eingelesen wird, der Abgleich mit Instrumenten
    erfolgt und die Preise aktualisiert werden.

    Parameter:
        settings (dict): Die Programmeinstellungen.
        instruments_df (DataFrame): DataFrame mit den Instrumenten-Daten.
        logfile (str): Der Name des Logfiles.
        screen (bool): Wenn True, werden Bildschirmmeldungen angezeigt.
        log (bool): Wenn True, werden Meldungen ins Logfile geschrieben.
        
    Rückgabe:
        DataFrame: Ein DataFrame mit den eingelesenen und verarbeiteten Preisdaten oder None bei Fehlern.
    """
    error_count = 0
    warning_count = 0
    prices_df = None

    try:
        # 1. Prices-Datei (Kurse) importieren
        prices_file = (settings or {}).get('Files', {}).get('prices', '')
        prices_df = import_parquet(prices_file, logger)
        if prices_df is None:
            logger.error(f"Fehler beim Einlesen der Kurse-Datei '{prices_file}'.")
            error_count += 1
            return None

        # 2. Abgleich von prices und instruments
        prices_check_for_instruments(prices_df, instruments_df, logger)

        # 3. Aktualisiere die Kurse in prices_df mit prices_update
        prices_df = prices_update(prices_df, instruments_df, logger)
        
        # 4. Speichere den aktualisierten DataFrame in die Parquet-Datei
        try:
            prices_df.to_parquet((settings or {}).get('Files', {}).get('prices', ''))
            logger.info(f"Aktualisierte Prices-Daten erfolgreich in Parquet-Datei '{(settings or {}).get('Files', {}).get('prices', '')}' gespeichert.")
        except Exception as e:
            logger.warning(f"Fehler beim Speichern der Prices-Parquet-Datei '{(settings or {}).get('Files', {}).get('prices', '')}': {e}")
            warning_count += 1

        # 5. Sicherstellen, dass alle (Datum, WKN)-Kombinationen vorhanden sind (ergänzt 26.3.25)
        # es werden alle Tage, auch wochenenden ergänzt (und später mit ffill aufgefüllt) so dass auch wenn ein Monatsende auf ein Wochenende fällt, der 
        # Monatsendwert korrekt berechnet wird
        try:

            today = datetime.today().date()
            yesterday = pd.Timestamp(today - timedelta(days=1))
            all_dates = pd.date_range(
                start=prices_df.index.get_level_values('date').min(),
                end=yesterday,
                freq='D'  # täglich – inkl. Wochenende
            )

            all_wkns = prices_df.index.get_level_values('wkn').unique()

            full_index = pd.MultiIndex.from_product([all_dates, all_wkns], names=['date', 'wkn'])

            # Nur 'price' beibehalten, Index auffüllen
            prices_df = prices_df.reindex(full_index)

            logger.info("Fehlende (Datum, WKN)-Kombinationen im DataFrame ergänzt mit NaN (alle Kalendertage).")

        except Exception as e:
            logger.warning(f"Fehler beim Ergänzen fehlender Datum-WKN-Kombinationen: {e}")
            warning_count += 1

    

        # 6. Fehlende Preise mit forward-fill pro WKN auffüllen (am 26.3.25 ergänzt)
        try:
            prices_df = prices_df.sort_index(level='date')  # Sicherstellen, dass nach Datum sortiert
            prices_df['price'] = prices_df.groupby('wkn')['price'].ffill()
            logger.info("Fehlende Preis Werte erfolgreich mittels ffill pro WKN aufgefüllt.")
        except Exception as e:
            logger.warning(f"Fehler beim Auffüllen der Preise mit ffill: {e}")
            warning_count += 1    

        # 7. Exportiere prices_df als Excel-Pivot-Datei, falls 'excel_pivot' in den Einstellungen vorhanden ist
        #
        # 23.11.24 ich habe die Abfrage deaktiviert, da es nach der umstellung der settings noch unklar ist, ob ich das über ein flag oder den filenamen aktiviere
        # wenn ich das sauber aufstelle, dann muss ich einen filenamen in der ini datei und ein flag definieren und bei der abfrage der verügbarkeit der files die abfrage
        # abhängig von dem flag
        #
        
        try:
            export_2D_df_to_excel_format(prices_df, (settings or {}).get("Export", {}).get("prices_to_excel", {}), logger)
        except Exception as e:
            logger.warning(f"Fehler beim Exportieren der Excel-Pivot-Datei '{(settings or {}).get("Export", {}).get("prices_to_excel", {}).get("filename", "")}': {e}")
            warning_count += 1

    except Exception as e:
        logger.error(f"Unerwarteter Fehler beim Import der Kurs-Daten: {e}")
        error_count += 1

    # Ergebnis ausgeben
    if error_count > 0:
        logger.error("Kursdaten-Import und -Verarbeitung fehlgeschlagen.")
    elif warning_count > 0:
        logger.warning(f"Kursdaten-Import und -Verarbeitung abgeschlossen mit {warning_count} Warnung(en).")
    else:
        logger.info("Kursdaten-Import und -Verarbeitung erfolgreich abgeschlossen.")
    return prices_df

# Main Block 04: Buchungen (bookings) importieren
def bookings_import_and_process(settings, instruments_df, logger):
    """
    Importiert und verarbeitet die Buchungen, indem die Bookings-Datei eingelesen und mit den Instrumenten abgeglichen wird.
    Überprüft, ob alle WKNs in den Buchungen auch in den Instrumenten vorhanden sind.

    Parameter:
        settings (dict): Die Programmeinstellungen.
        instruments_df (DataFrame): DataFrame mit den Instrumenten-Daten.
        logfile (str): Der Name des Logfiles.
        screen (bool): Wenn True, werden Bildschirmmeldungen angezeigt.
        log (bool): Wenn True, werden Meldungen ins Logfile geschrieben.
        
    Rückgabe:
        bookings_df (DataFrame): Ein DataFrame mit den eingelesenen und verarbeiteten Buchungsdaten oder None bei Fehlern.
    """
    error_count = 0
    warning_count = 0
    bookings_df = None

    try:
        # 1 Bookings-Datei (Buchungen) importieren
        bookings_file = (settings or {}).get('Files', {}).get('bookings', '')
        bookings_df = bookings_import(bookings_file, logger)
        
        if bookings_df is None:
            logger.error(f"Fehler beim Import der Buchungs-Datei '{bookings_file}'.")
            error_count += 1
            return None

        # 2 Prüfe bookings_df gegen instruments_df und breche ab, wenn WKNs fehlen
        missing_wkns = bookings_check_for_instruments(bookings_df, instruments_df)
        
        if missing_wkns:
            logger.error(f"Die folgenden WKNs aus 'bookings_df' fehlen in 'instruments_df': {missing_wkns}")
            error_count += 1
            return None
        else:
            logger.info("Alle WKNs aus 'bookings_df' sind in 'instruments_df' vorhanden.")

    except Exception as e:
        logger.error(f"Unerwarteter Fehler beim Einlesen der Buchungen: {e}")
        error_count += 1

    # Ergebnis ausgeben
    if error_count > 0:
        logger.error("Buchungen-Import und -Verarbeitung fehlgeschlagen.")
    elif warning_count > 0:
        logger.warning(f"Buchungen-Import und -Verarbeitung abgeschlossen mit {warning_count} Warnung(en).")
    else:
        logger.info("Buchungen-Import und -Verarbeitung erfolgreich abgeschlossen.")
    return bookings_df



# Main Block 09: Portofolio Wert nach Anlagetypen bzw Regionen
def export_portfolio_analysis(values_day_df, instruments_type_df, instruments_region_df, logger):

    values_month_df = df_to_eom(values_day_df)
    
    # 9.1.
    values_type_month_df = values_type_month(values_month_df, instruments_type_df)
    export_2D_df_to_excel_format(values_type_month_df, (settings or {}).get("Export", {}).get("values_type_month_to_excel", {}), logger)

    # 9.2. Portfolio Zusammensetzung - prozentualer Anteil pro Anlagetyp
    values_type_month_percentage_df = df_transform_each_line_to_percentage(values_type_month_df)
    export_2D_df_to_excel_format(values_type_month_percentage_df, (settings or {}).get("Export", {}).get("values_type_month_percentage_to_excel", {}), logger)

    # 9.3. Portfolio Wert nach Regionen
    values_region_month_df = values_region_month(values_month_df, instruments_region_df, logger)
    export_2D_df_to_excel_format(values_region_month_df, (settings or {}).get("Export", {}).get("values_region_month_to_excel", {}), logger)

    # 9.4. Portfolio Zusammensetzung - prozentualer Anteil pro Region
    values_region_month_percentage_df = df_transform_each_line_to_percentage(values_region_month_df)
    export_2D_df_to_excel_format(values_region_month_percentage_df, (settings or {}).get("Export", {}).get("values_region_month_percentage_to_excel", {}), logger)

    # 9.5. nur Instrumente, die einen Regional Charakter haben 
    values_region_month_wo_exception_df = values_region_month_df.drop(index='exception', level='reg')
    values_region_month_wo_exception_percentage_df = df_transform_each_line_to_percentage(values_region_month_wo_exception_df)
    export_2D_df_to_excel_format(values_region_month_wo_exception_percentage_df, (settings or {}).get("Export", {}).get("values_region_month_wo_exception_percentage_to_excel", {}), logger)

    return





# Holt die Werte für Cash Rückstellungen aus Provisions.xlsx
def provisions_month_import_and_process(values_month_df, settings, logger):
    """
    Verarbeitet den DataFrame 'provisions_month_df', indem die Datumswerte aus 'values_month_df' übernommen werden,
    basierend auf dem Abgleich von Monat und Jahr. Fehlende Datumswerte werden ergänzt und Lücken gefüllt.

    Parameter:
        values_month_df (DataFrame): DataFrame mit einem MultiIndex aus 'date' und 'wkn'.
        settings (dict): Programmeinstellungen mit dem Dateipfad für Provisionsdaten.
        logfile (str): Name des Logfiles.
        screen (bool): Ob Ausgaben auf dem Bildschirm erfolgen sollen.
        log (bool): Ob Ausgaben ins Logfile geschrieben werden sollen.

    Rückgabe:
        DataFrame: Der verarbeitete DataFrame 'provisions_month_df' mit angepassten Datumswerten und gefüllten Lücken.
    """

    error_count = 0
    warning_count = 0
    provisions_month_df = None

    try:
        # Provisionsdatei aus den Einstellungen laden
        provisions_file = (settings or {}).get('Files', {}).get('provisions', '')
        if not provisions_file:
            raise FileNotFoundError("Provisions-Dateipfad nicht in den Einstellungen gefunden.")

        # Einlesen der Excel-Datei mit zwei Spalten: Datum und Wert
        provisions_month_df = pd.read_excel(provisions_file, usecols=[0, 1], names=['date', 'provision'])
        provisions_month_df['date'] = pd.to_datetime(provisions_month_df['date'])
        provisions_month_df.set_index('date', inplace=True)

        # Erstelle eine Liste mit eindeutigen Datumswerten aus values_month_df (nur der Date-Part)
        unique_dates = values_month_df.index.get_level_values('date').unique()

        # Erstelle eine neue Spalte 'new_date' in provisions_month_df und ersetze basierend auf Monat/Jahr
        provisions_month_df['new_date'] = provisions_month_df.index
        for provision_date in provisions_month_df.index:
            for unique_date in unique_dates:
                if (provision_date.month == unique_date.month) and (provision_date.year == unique_date.year):
                    provisions_month_df.at[provision_date, 'new_date'] = unique_date
                    break
       
        # Setze den neuen 'date'-Index
        provisions_month_df.reset_index(drop=True, inplace=True)
        provisions_month_df.set_index('new_date', inplace=True)
        provisions_month_df.index.name = 'date'
              
        # Ergänze fehlende Datumswerte aus values_month_df
        provisions_month_df = provisions_month_df.reindex(unique_dates, fill_value=np.nan)

        # Führe ein 'ffill' aus, um fehlende Werte aufzufüllen
        provisions_month_df['provision'] = provisions_month_df['provision'].ffill()

        # Überprüfe, ob der letzte Wert immer noch null oder NaN ist und ersetze ihn ggf. durch den vorletzten Wert
        if pd.isnull(provisions_month_df['provision'].iloc[-1]) or provisions_month_df['provision'].iloc[-1] == 0:
            provisions_month_df['provision'].iloc[-1] = provisions_month_df['provision'].iloc[-2]

        logger.info("Provisions Daten erfolgreich importiert und verarbeitet.")

    except FileNotFoundError as e:
        logger.error(f"{e}")
        error_count += 1
        provisions_month_df = None
    except Exception as e:
        logger.error(f"Unerwarteter Fehler beim Import und Verarbeiten der Provisionsdaten: {e}")
        error_count += 1
        provisions_month_df = None

    # Ergebnis ausgeben
    if error_count > 0:
        logger.error("Provisions-Import und -Verarbeitung fehlgeschlagen.")
    elif warning_count > 0:
        logger.warning(f"Provisions-Import und -Verarbeitung abgeschlossen mit {warning_count} Warnung(en).")
    else:
        logger.info("Provisions-Import und -Verarbeitung erfolgreich abgeschlossen.")

    return provisions_month_df

# Erstellt angepaßten Values Dataframe, bei dem der Cash Anteil um die Provisions reduziert ist
def values_month_adjust_for_provisions(values_month_df, provisions_month_df, logger):
    """
    Passt den DataFrame 'values_month_df' basierend auf den Provisionswerten in 'provisions_month_df' an.

    Parameter:
        values_month_df (DataFrame): DataFrame mit Monatsdaten.
        provisions_month_df (DataFrame): DataFrame mit Provisionsdaten.
        logger (ExtendedLogger): Logger instance for output and logging.

    Rückgabe:
        DataFrame: Ein angepasster DataFrame, in dem die Werte für WKN 'cash' basierend auf den Provisionswerten angepasst wurden.
    """
    # Kopiere den ursprünglichen DataFrame, um Änderungen vorzunehmen
    values_adjusted_df = values_month_df.copy()

    # Iteriere über alle Datumswerte in values_month_df
    for date in values_month_df.index.get_level_values('date').unique():
        # Überprüfe, ob das Datum in provisions_month_df vorhanden ist
        if date in provisions_month_df.index:
            provision_value = provisions_month_df.loc[date, 'provision']
            
            # Überprüfe, ob WKN 'cash' für das aktuelle Datum in values_month_df vorhanden ist
            if ('cash' in values_month_df.loc[date].index):
                cash_value = values_month_df.loc[(date, 'cash'), 'value']
                
                if cash_value > provision_value:
                    # Reduziere den Wert um die Provision
                    values_adjusted_df.loc[(date, 'cash'), 'value'] -= provision_value
                else:
                    # Gebe eine WARNING aus und setze den Cash-Wert auf 0
                    warning_message = (f"Der Cash-Wert am {date} ist kleiner oder gleich der Provision. "
                                       f"Der Wert wurde auf 0 gesetzt.")
                    logger.warning(warning_message)
                    values_adjusted_df.loc[(date, 'cash'), 'value'] = 0

    return values_adjusted_df

# Erstellt angepaßten Values nach Anlagetyp Dataframe, bei dem der Cash Anteil um die Provisions reduziert ist
def values_type_month_after_provisions(values_type_month_df, provisions_month_df, logger):
    """
    Passt die Werte im DataFrame `values_type_month_df` an, indem der Wert für `cash`
    um die entsprechenden Werte in `provisions_month_df` reduziert wird.

    Parameter:
        values_type_month_df (DataFrame): Ein DataFrame mit monatlichen Werten pro Typ (MultiIndex: date, type).
        provisions_month_df (DataFrame): Ein DataFrame mit den monatlichen Provisionswerten (Index: date).
        logfile (str): Name des Logfiles.
        screen (bool): Ob Ausgaben auf dem Bildschirm erfolgen sollen.

    Rückgabe:
        DataFrame | None: Der angepasste DataFrame mit umbenannter Spalte `cash` zu `cash_invest`,
        oder None bei Fehlern.
    """
    try:
        # Überprüfen, ob die erforderliche Spalte 'cash' im DataFrame vorhanden ist
        if 'cash' not in values_type_month_df.index.get_level_values('type'):
            raise KeyError("Die Spalte 'cash' fehlt im DataFrame 'values_type_month_df'.")

        # Überprüfen, ob alle Datumswerte in provisions_month_df im Index von values_type_month_df enthalten sind
        if not provisions_month_df.index.isin(values_type_month_df.index.get_level_values('date')).all():
            raise ValueError("Die Index-Daten in 'provisions_month_df' stimmen nicht mit 'values_type_month_df' überein.")

        # Kopiere den ursprünglichen DataFrame
        adjusted_df = values_type_month_df.copy()

        # Iteriere über alle Datumswerte in provisions_month_df
        for date in provisions_month_df.index:
            # Überprüfen, ob der Wert für 'cash' existiert
            if ('cash' in adjusted_df.loc[date].index):
                provision_value = provisions_month_df.loc[date, 'provision']
                adjusted_df.loc[(date, 'cash'), 'type_value'] -= provision_value

                # Überprüfen, ob der neue Wert negativ ist und eine Warnung ausgeben
                if adjusted_df.loc[(date, 'cash'), 'type_value'] < 0:
                    adjusted_df.loc[(date, 'cash'), 'type_value'] = 0
                
                    warning_message = (f"Der Wert von 'cash' am {date} wurde auf 0 gesetzt, da er kleiner als die Provision war.")
                    logger.warning(warning_message)

        # Benenne die Spalte 'cash' in 'cash_invest' um
        adjusted_df = adjusted_df.rename(index={'cash': 'cash_invest'})

        # Rückgabe des angepassten DataFrames
        return adjusted_df

    except Exception as e:
        error_message = f"ERROR: Ein Fehler ist in 'values_type_month_after_provisions' aufgetreten: {e}"
        logger.error(error_message)
        return None

# Holt die Werte für die Zielstruktur des Portfolios aus Bookings.xlsx
def target_shares_import_and_process(instruments_filename, logger):
    """
    Importiert das Excel-File und liest die Spalten "wkn" und "Ziel" ein.
    Überprüft, ob die Summe von "target_share" 100% beträgt.

    Parameter:
        instruments_filename (str): Der Dateiname des Excel-Files.
        logfile (str): Name des Logfiles.
        screen (bool): Ob Ausgaben auf dem Bildschirm erfolgen sollen.
        log (bool): Ob Ausgaben ins Logfile geschrieben werden sollen.

    Rückgabe:
        DataFrame: DataFrame mit den Spalten "wkn" und "target_share", oder ein leerer DataFrame bei Fehler.
    """
    error_count = 0
    warning_count = 0
    target_df = pd.DataFrame()

    try:
        # Excel-Datei lesen
        df = pd.read_excel(instruments_filename, usecols=["wkn", "Ziel"])
        
        # Spalten umbenennen und "wkn" als Index setzen
        df.rename(columns={"Ziel": "target_share"}, inplace=True)
        df['wkn'] = df['wkn'].str.lower()
        df.set_index("wkn", inplace=True)
        
        # Überprüfung der Summe der "target_share"
        total_target_share = round(df['target_share'].sum(), 1)
        if total_target_share == 1:
            target_df = df
            logger.info("Target-Daten erfolgreich importiert und überprüft.")
        else:
            logger.warning(f"Die Summe der 'target_share' beträgt {total_target_share*100}% und ist nicht 100%.")
            warning_count += 1

    except FileNotFoundError as e:
        logger.error(f"Datei '{instruments_filename}' nicht gefunden: {e}")
        error_count += 1
    except Exception as e:
        logger.error(f"Unerwarteter Fehler beim Import der Target-Daten: {e}")
        error_count += 1

    # Ergebnis ausgeben
    if error_count > 0:
        logger.error("Import der Target-Daten fehlgeschlagen.")
    elif warning_count > 0:
        logger.warning(f"Import der Target-Daten abgeschlossen mit {warning_count} Warnung(en).")
    else:
        logger.info("Import der Target-Daten erfolgreich abgeschlossen.")

    if warning_count > 0:
        target_df=None

    return target_df

def values_vs_target(values_month_df, target_shares_df, prices_df, logger):
    """
    Vergleicht die aktuellen Kontostände mit den Zielanteilen und berechnet die Abweichungen.

    Parameter:
        values_month_df (DataFrame): DataFrame mit monatlichen Kontoständen.
        target_shares_df (DataFrame): DataFrame mit Zielanteilen der WKNs.
        prices_df (DataFrame): DataFrame mit aktuellen Preisen.
        logfile (str): Name des Logfiles.
        screen (bool): Ob Ausgaben auf dem Bildschirm erfolgen sollen.
        log (bool): Ob Ausgaben ins Logfile geschrieben werden sollen.

    Rückgabe:
        DataFrame: Der DataFrame 'values_deviation_from_target_df'.
    """
    error_count = 0
    warning_count = 0

    try:
        # Extrahiere den jüngsten Eintrag aus values_month_df
        latest_date = values_month_df.index.get_level_values('date').max()
        values_actual_df = values_month_df.xs(latest_date, level='date')
        if (settings or {}).get('Output', {}).get('debug', False): export_df_to_excel(values_actual_df, "values_actual_debug.xlsx", logger)
        
        # Berechne den total_value als Summe über alle WKNs
        total_value = values_actual_df['value'].sum()
        if (settings or {}).get('Output', {}).get('debug', False): export_df_to_excel(target_shares_df, "target_shares_debug.xlsx", logger)

        # Sicherstellen, dass der Index von target_shares_df dem von values_actual_df entspricht
        target_shares_df = target_shares_df.reindex(values_actual_df.index)
        if (settings or {}).get('Output', {}).get('debug', False): export_df_to_excel(target_shares_df, "target_shares_debug.xlsx", logger)
                
        # Erstelle den values_target_df
        values_target_df = values_actual_df.copy()
        values_target_df['value'] = total_value * target_shares_df['target_share']
        if (settings or {}).get('Output', {}).get('debug', False): export_df_to_excel(values_target_df, "values_target_debug.xlsx", logger)
            
        # Erstelle values_delta_df durch Subtraktion
        values_delta_df = values_actual_df['value'] - values_target_df['value']
        values_delta_df = values_delta_df.to_frame(name='delta')
        if (settings or {}).get('Output', {}).get('debug', False): export_df_to_excel(values_delta_df, "values_delta_debug.xlsx", logger)
        
        # Berechne den aktuellen Kurs für das jüngste Datum aus prices_df
        latest_prices_df = prices_df.xs(latest_date, level='date')
                
        # Vermeide Division durch 0, indem 0-Werte in 'price' durch NaN ersetzt werden (Division durch NaN ergibt NaN)
        latest_prices_df = latest_prices_df.copy()
        latest_prices_df['price'] = latest_prices_df['price'].replace(0, np.nan)
        if (settings or {}).get('Output', {}).get('debug', False): export_df_to_excel(latest_prices_df, "latest_prices_debug.xlsx", logger)
        
        # Berechne buy_sell_df
        buy_sell_df = values_delta_df['delta'] / latest_prices_df['price']
        buy_sell_df = buy_sell_df.to_frame(name='buy_sell')
        if (settings or {}).get('Output', {}).get('debug', False): export_df_to_excel(buy_sell_df, "buy_sell_debug.xlsx", logger)
        
        # Ergebnis-Log und Rückgabe
        logger.info("Abweichungen von Zielporfolio erfolgreich berechnet.")
    
    except Exception as e:
        logger.error(f"Unerwarteter Fehler bei der Berechnung der Abweichungen: {e}")
        error_count += 1
        buy_sell_df = pd.DataFrame()  # Leerer DataFrame als Fallback

    # Ergebnis ausgeben
    if error_count > 0:
        logger.error("Kontostände vs. Zielberechnung fehlgeschlagen.")
    elif warning_count > 0:
        logger.warning(f"Kontostände vs. Zielberechnung abgeschlossen mit {warning_count} Warnung(en).")
    else:
        logger.info("Kontostände vs. Zielberechnung erfolgreich abgeschlossen.")

    return buy_sell_df

def overview(values_df, unrealized_gains_losses_df, invest_df, logger):
    """
    Erstellt eine Übersicht über die Summen pro Datumseintrag der "non_cash_instruments" und "cash_instruments",
    sowie (ohne Berechnung) den Gewinn/Verlust und Investitionswerte.

    Parameter:
        values_df (DataFrame): DataFrame mit Kontoständen (pro periode) (MultiIndex: date, wkn).
        unrealized_gains_losses_df (DataFrame): DataFrame mit Gewinn-Verlust-Werten (pro periode) (MultiIndex: date, wkn).
        invest_df (DataFrame): DataFrame mit täglichen Investitionswerten (pro periode) (Index: date).
        logger (ExtendedLogger): Logger instance for output and logging.

    Rückgabe:
        DataFrame: Ein DataFrame mit Index Datum mit den Spalten 'non_cash_instruments', 'cash_instruments', 'unrealized_gains_losses', 'invest'.
    """
    try:
        # Berechnung der Summen für non_cash_instruments und cash_instruments
        non_cash_mask = ~values_df.index.get_level_values('wkn').isin(['cash', 'crypto', 'cm', 'ftd'])
        cash_mask = values_df.index.get_level_values('wkn').isin(['cash', 'crypto', 'cm', 'ftd'])

        non_cash_instruments = values_df[non_cash_mask].groupby('date')['value'].sum()
        cash_instruments = values_df[cash_mask].groupby('date')['value'].sum()

        # Gewinn/Verlust summieren pro Tag
        unrealized_gains_losses = unrealized_gains_losses_df.groupby('date')['unrealized_gains_losses'].sum()

        # Investitionswerte
        invest = invest_df['invest']

        # Zusammenführen der Daten
        overview_df = pd.DataFrame({
            'non_cash_instruments': non_cash_instruments,
            'cash_instruments': cash_instruments,
            'unrealized_gains_losses': unrealized_gains_losses,
            'invest': invest
        })

        # Fehlende Werte mit 0 auffüllen
        overview_df.fillna(0, inplace=True)

        # Log-Erfolgsmeldung
        logger.info("Info: Übersicht erfolgreich erstellt.")

        return overview_df

    except Exception as e:
        logger.error(f"Fehler bei der Erstellung der Übersicht: {e}")
        return pd.DataFrame()

def export_bank_analysis_to_excel(df_shares, df_values, filename, logger):
    """
    Erstellt eine Excel-Datei mit Analysen pro Bank. Für jede Bank wird eine eigene Matrix erstellt,
    wobei die Zeilen das Datum und die Spalten WKN-bezogene Daten (Shares und Values) sind.
    Vor dem Export werden Datumszeilen, in denen alle Werte 0 sind, und WKN-Spaltenpaare,
    in denen alle 'share'-Werte 0 sind, entfernt.
    Die Spalten werden in der Reihenfolge wkn1-share, wkn1-value, wkn2-share, wkn2-value sortiert.

    Banken ohne verbleibende Daten nach der Bereinigung werden nicht exportiert.

    Parameter:
        df_shares (DataFrame): DataFrame mit MultiIndex (date, bank, wkn) und einer Spalte für die Anteile (share).
        df_values (DataFrame): DataFrame mit MultiIndex (date, bank, wkn) und einer Spalte für die Werte (value).
        filename (str): Name der Excel-Datei, die erzeugt werden soll.
        logger (ExtendedLogger): Logger instance for output and logging.
    """
    try:
        # Initialisiere die Excel-Arbeitsmappe
        workbook = Workbook()
        workbook.remove(workbook.active)  # Entferne das Standard-Arbeitsblatt

        # Überprüfen, ob die erwarteten Spalten vorhanden sind
        if 'share' not in df_shares.columns:
            raise ValueError("Die Spalte 'share' fehlt im DataFrame 'df_shares'.")
        if 'value' not in df_values.columns:
            raise ValueError("Die Spalte 'value' fehlt im DataFrame 'df_values'.")

        # Überprüfe, ob die beiden DataFrames denselben Index haben
        if not df_shares.index.equals(df_values.index):
            raise ValueError("Die Indizes von df_shares und df_values stimmen nicht überein.")

        # Liste der einzigartigen Banken
        banks = df_shares.index.get_level_values('bank').unique()

        for bank in banks:
            # Filtere die Daten für die aktuelle Bank
            shares_bank = df_shares.xs(bank, level='bank')
            values_bank = df_values.xs(bank, level='bank')

            # Erstelle die Matrix
            combined = pd.concat(
                [shares_bank['share'], values_bank['value']],
                axis=1,
                keys=['share', 'value']
            )
            combined = combined.unstack(level='wkn')

            # Bereinigung: Entferne WKN-Spalten mit nur 0 in 'share'
            non_zero_wkns = combined['share'].columns[combined['share'].sum(axis=0) != 0]
            combined = combined.loc[:, (slice(None), non_zero_wkns)]

            # Bereinigung: Entferne Datums-Zeilen mit nur 0-Werten in allen Spalten
            combined = combined.loc[~(combined.fillna(0).sum(axis=1) == 0)]

            # Überspringe Banken ohne verbleibende Daten
            if combined.empty:
                logger.info(f"Keine Daten vorhanden für Bank '{bank}'. Überspringe Export.")
                continue

            # Umsortieren der Spalten: wkn1-share, wkn1-value, wkn2-share, wkn2-value
            new_column_order = []
            for wkn in non_zero_wkns:
                new_column_order.append(('share', wkn))
                new_column_order.append(('value', wkn))
            combined = combined.loc[:, new_column_order]

            # Erstelle ein Arbeitsblatt für die aktuelle Bank
            sheet = workbook.create_sheet(title=f"Bank_{bank}")

            # Füge die bereinigte und sortierte Matrix in das Arbeitsblatt ein
            for row in dataframe_to_rows(combined, index=True, header=True):
                sheet.append(row)

        # Speichere die Excel-Datei, wenn mindestens ein Arbeitsblatt erstellt wurde
        if len(workbook.sheetnames) > 0:
            workbook.save(filename)
            logger.info(f"Datei '{filename}' erfolgreich erstellt.")
        else:
            logger.info(f"Keine Daten vorhanden für irgendeine Bank. Datei '{filename}' wurde nicht erstellt.")

    except Exception as e:
        logger.error(f"Fehler beim Erstellen der Excel-Datei '{filename}': {e}")

# Erstellt Übersicht Non-Cash und Cash pro Bank für den Export in das Finance File
def depots_fuer_finance(values_month_banks_df, logger):
    """
    Erstellt eine Auswertung der Cash-Werte und Nicht-Cash-Werte pro Bank auf Basis von values_month_banks_df.

    Parameter:
        values_month_banks_df (DataFrame): MultiIndex (date, bank, wkn) mit Spalte 'value'.
        logger (ExtendedLogger): Logger instance for output and logging.

    Rückgabe:
        DataFrame: Index = Datum, Spalten = Bank_cash, Bank_non_cash
    """
    try:
        if not isinstance(values_month_banks_df.index, pd.MultiIndex):
            raise ValueError("Der DataFrame muss einen MultiIndex mit (date, bank, wkn) besitzen.")

        # Cash und Nicht-Cash trennen
        is_cash = values_month_banks_df.index.get_level_values("wkn") == "cash"
        df_cash = values_month_banks_df[is_cash].copy()
        df_non_cash = values_month_banks_df[~is_cash].copy()

        # Gruppieren und pivotieren
        df_cash_grouped = df_cash.groupby(["date", "bank"]).sum().unstack(fill_value=0)
        df_non_cash_grouped = df_non_cash.groupby(["date", "bank"]).sum().unstack(fill_value=0)

        # Umbenennen der Spalten
        df_cash_grouped.columns = [f"{bank}_cash" for bank in df_cash_grouped.columns.get_level_values(1)]
        df_non_cash_grouped.columns = [f"{bank}_non_cash" for bank in df_non_cash_grouped.columns.get_level_values(1)]

        # Zusammenführen
        result_df = pd.concat([df_cash_grouped, df_non_cash_grouped], axis=1)
        result_df = result_df.sort_index(axis=1)

        logger.info("depots_fuer_finance erfolgreich erstellt.")
        return result_df

    except Exception as e:
        logger.error(f"Fehler in depots_fuer_finance: {e}")
        return None

def export_overview(values_day_df, unrealized_gains_losses_day_df, invest_day_df, logger):
    """ Overview (Bericht)
        erstellt die overviews und exportiert diese
        keine rückgabe daten
    

    """

    format_numbers=["DD.MM.YY","#,##0 €","#,##0 €","#,##0 €;[Red]-#,##0 €","#,##0 €;[Red]-#,##0 €"]
    format_columns=[9, 12, 10, 10, 10]

    overview_day_df = overview(values_day_df, unrealized_gains_losses_day_df, invest_day_df, logger)
    if (settings or {}).get("Export", {}).get("overview_day_to_excel", {}).get("enabled", False):
        export_df_to_excel(overview_day_df, (settings or {}).get("Export", {}).get("overview_day_to_excel", {}).get("filename", ""), logger)

    values_month_df = df_to_eom(values_day_df)
    unrealized_gains_losses_month_df = df_2D_sum_per_period(unrealized_gains_losses_day_df, 'month')
    invest_month_df = df_1D_sum_per_period(invest_day_df, 'month')

    overview_month_df = overview(values_month_df, unrealized_gains_losses_month_df, invest_month_df, logger)
    if (settings or {}).get("Export", {}).get("overview_month_to_excel", {}).get("enabled", False):
        export_df_to_excel(overview_month_df, (settings or {}).get("Export", {}).get("overview_month_to_excel", {}).get("filename", ""), logger)
    if (settings or {}).get("Export", {}).get("overview_month_to_excel", {}).get("enabled", False):
        format_excel_as_table_with_freeze((settings or {}).get("Export", {}).get("overview_month_to_excel", {}).get("filename", ""), logger, table_name="Table1", style_name="TableStyleMedium1", freeze_first_row=True)
        format_excel_columns((settings or {}).get("Export", {}).get("overview_month_to_excel", {}).get("filename", ""), format_numbers, logger, format_columns)
    
    
    values_year_df = df_to_eoy(values_day_df)
    unrealized_gains_losses_year_df = df_2D_sum_per_period(unrealized_gains_losses_day_df, 'year')
    invest_year_df = df_1D_sum_per_period(invest_day_df, 'year')

    overview_year_df = overview(values_year_df, unrealized_gains_losses_year_df, invest_year_df, logger)
    if (settings or {}).get("Export", {}).get("overview_year_to_excel", {}).get("enabled", False):
        export_df_to_excel(overview_year_df, (settings or {}).get("Export", {}).get("overview_year_to_excel", {}).get("filename", ""), logger)
    if (settings or {}).get("Export", {}).get("overview_year_to_excel", {}).get("enabled", False):
        format_excel_as_table_with_freeze((settings or {}).get("Export", {}).get("overview_year_to_excel", {}).get("filename", ""), logger, table_name="Table1", style_name="TableStyleMedium1", freeze_first_row=True)
        format_excel_columns((settings or {}).get("Export", {}).get("overview_year_to_excel", {}).get("filename", ""), format_numbers, logger, format_columns)

    return





# Hauptprogramm
if __name__ == "__main__":
    # 1. Initialisierung
    settings = initializing('depot.ini', screen=False)
    logfile=(settings or {}).get('Files', {}).get('logfile', '')
    screen=(settings or {}).get('Output', {}).get('screen', False)

    # Initialize ahlib logger
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    logger = create_extended_logger(logfile, screen, script_name=script_name)
    logger.info(f"Programm {script_name} wird gestartet")

    # 2. Instruments
    instruments_df, instruments_region_df, instruments_type_df = instruments_import_and_process(settings, logger)

    # 3. Prices-Datei (Kurse) importieren, verarbeiten und überwachen
    prices_df = prices_import_and_process(settings, instruments_df, logger)
    if prices_df is None:
        logger.error("Prices-Datei konnte nicht geladen werden. Beende das Programm.")
        exit(1)
    end_date = prices_df.index.get_level_values('date').max()
    start_date = prices_df.index.get_level_values('date').min()

    # 4. Bookings-Datei (Buchungen)
    bookings_df = bookings_import_and_process(settings, instruments_df, logger)

    # 5. Shares (Bestand) auf Tagesbasis pro wkn und Bank aus bookings_df (Buchungen) aufgebaut
    shares_day_banks_df = shares_from_bookings(bookings_df, end_date, logger)
    

    # 6. Values (Depot-/Kontostände in Euro, Stück*Preis = Wert)
    # Benötigte Daten: shares_day_banks_df, prices_df

    # 6.1. Values (Kontostände in Euro) aus Positions (Bestände in Stück) und prices_df (Kurse) aufgebaut
    values_day_banks_df = values_from_shares_and_prices(shares_day_banks_df, prices_df)
    logger.info('Werte (values) erfolgreich aufgebaut')

    # 6.2. Reduziere values_day_banks_df auf die Monatsebene
    values_month_banks_df = df_to_eom(values_day_banks_df)
    logger.info('Werte (values) daily erfolgreich auf Monatsebene reduziert')
    if (settings or {}).get("Export", {}).get("values_month_banks_to_excel", {}).get("enabled", False):
        export_df_to_excel(values_month_banks_df, (settings or {}).get("Export", {}).get("values_month_banks_to_excel", {}).get("filename", ""), logger)

    # 6.3. Aggregiere die Werte in values_month_banks_df über alle Banken
    values_month_df = aggregate_banks(values_month_banks_df)
    logger.info('Werte (values) auf Monatsebene über Banken erfolgreich aggregiert')
    export_2D_df_to_excel_format(values_month_df, (settings or {}).get("Export", {}).get("values_month_to_excel", {}), logger)

    # 6.4. Aggregiere die Werte in values_day_banks_df über alle Banken
    values_day_df = aggregate_banks(values_day_banks_df)
    logger.info('Werte (values) auf Tagesebene über Banken erfolgreich aggregiert')
    export_2D_df_to_excel_format(values_day_df, (settings or {}).get("Export", {}).get("values_day_to_excel", {}), logger)
    
    # 6.5. Reduziere values_day_banks_df auf die Jahresebene (Ende jedes Jahres oder letztes verfügbares Datum)
    #values_year_banks_df = df_to_eoy(values_day_banks_df)
    #screen_and_log('Info: Werte (values) daily erfolgreich auf Jahresebene reduziert', logfile, screen=screen)

    # 6.6. Aggregiere die Werte in values_year_banks_df über alle Banken
    #values_year_df = aggregate_banks(values_year_banks_df)
    #screen_and_log('Info: Werte (values) auf Jahresebene über Banken erfolgreich aggregiert', logfile, screen=screen)
    #if settings['Export']['values_year_to_excel']: export_df_to_excel(values_year_df, "values_year_export.xlsx", logger)
    #export_2D_df_to_excel_pivot(values_year_df, "values_year_pivot_export.xlsx", logfile, screen=False)
   

    # 7. Gewinn und Verlust (über alle Banke aggregiert) aus Kursentwicklung (keine Gebühren, Steuern, Zinsen betrachtet) und Veränderung in Cash
    # Benötigte Daten: shares_day_banks_df bzw shares_day_df, prices_df, values_day_df, settings

    # 7.1. Bestimme Buch-Gewinne und Verluste auf Tagesbasis (Anzahl des jeweiligen Instrumemts * Kurs-Differenz) vergleichbar Laspeyers 
    shares_day_df = aggregate_banks(shares_day_banks_df)
    export_2D_df_to_excel_format(shares_day_df, (settings or {}).get("Export", {}).get("shares_day_to_excel", {}), logger)

    unrealized_gains_losses_day_df = unrealized_gains_losses_day(shares_day_df, prices_df)
    export_2D_df_to_excel_format(unrealized_gains_losses_day_df, (settings or {}).get("Export", {}).get("unrealized_gains_losses_day_to_excel", {}), logger)
    

    # 7.2. Taxes and Fees aus bookings.xlsx holen
    bookings_filename = (settings or {}).get('Files', {}).get('bookings', '')

    fees_bank_df=fees_import(bookings_filename, logger)
    if (settings or {}).get("Export", {}).get("fees_bank_to_excel", {}).get("enabled", False):
        export_df_to_excel(fees_bank_df, (settings or {}).get("Export", {}).get("fees_bank_to_excel", {}).get("filename", ""), logger)
    fees_df=aggregate_banks(fees_bank_df)
    export_2D_df_to_excel_format(fees_df, (settings or {}).get("Export", {}).get("fees_to_excel", {}), logger)
        
    taxes_bank_df=taxes_import(bookings_filename, logger)
    if (settings or {}).get("Export", {}).get("taxes_bank_to_excel", {}).get("enabled", False):
        export_df_to_excel(taxes_bank_df, (settings or {}).get("Export", {}).get("taxes_bank_to_excel", {}).get("filename", ""), logger)
    taxes_df=aggregate_banks(taxes_bank_df)
    
   
    # 7.3. Interests and Dividends aus bookings.xlsx holen
    bookings_filename = (settings or {}).get('Files', {}).get('bookings', '')
    interest_dividends_bank_df=interest_dividends_import(bookings_filename, logger)
    if (settings or {}).get("Export", {}).get("interest_dividends_bank_to_excel", {}).get("enabled", False):
        export_df_to_excel(interest_dividends_bank_df, (settings or {}).get("Export", {}).get("interest_dividends_bank_to_excel", {}).get("filename", ""), logger)
    interest_dividends_df=aggregate_banks(interest_dividends_bank_df)

    # 7.3. Transaction_value_at_price (Käufe und Verkäufe zum Kurswert)
    bookings_filename = (settings or {}).get('Files', {}).get('bookings', '')
    transaction_value_at_price_bank_day_df=transaction_value_at_price_import(bookings_filename, logger)
    transaction_value_at_price_day_df=aggregate_banks(transaction_value_at_price_bank_day_df)
    export_2D_df_to_excel_format(transaction_value_at_price_day_df, (settings or {}).get("Export", {}).get("transaction_value_at_price_day_to_excel", {}), logger)

    # 7.4. Realized Gains and Loses - Summe aus Gebühren, Steuern und Zinsen / Dividenden die einer WKN zugeordnet sind
    # Der Dataframe ist tageweise aufgebaut, enthält nur Daten an denen ein relevante Buchung auftritt
    # Dataframe hat Multiindex (Datum, WKN) und eine Wertspalte
    realized_gains_losses_day_df = realized_gains_losses_day(fees_df, taxes_df, interest_dividends_df)
    export_2D_df_to_excel_format(realized_gains_losses_day_df, (settings or {}).get("Export", {}).get("realized_gains_losses_day_to_excel", {}), logger)

    # 7.5. Berechnet Buchgewinne pro Tag aus den Tageswerten und den Transaktionswerten ohne Gebühren und Steuern (Stück Kauf/Verkauf * Kurs)
    gains_losses_before_fees_taxes_day_df = gains_losses_before_fees_taxes_day(values_day_df, transaction_value_at_price_day_df)
    export_2D_df_to_excel_format(gains_losses_before_fees_taxes_day_df, (settings or {}).get("Export", {}).get("gains_losses_before_fees_taxes_day_to_excel", {}), logger)

    # 7.6. Buchgewinne und mit Gebühren, Steuern, Zinsen/Dividenden zusammenführen
    # gains_losses_after_fees_taxes_day_df

    # 7.7. Berechne tägliche Profitabilität pro WKN und Tag
    #yield_excl_div_day_df = yield_day_from_values_day(unrealized_gains_losses_day_df, values_day_df)
    yield_excl_div_day_df = yield_day_from_values_day(gains_losses_before_fees_taxes_day_df, values_day_df, logger)
    if yield_excl_div_day_df is not None:
        export_2D_df_to_excel_format(yield_excl_div_day_df, (settings or {}).get("Export", {}).get("yield_excl_div_day_to_excel", {}), logger)
    else:
        logger.warning("Tägliche Profitabilität (Yield) konnte nicht erstellt werden.")

    # 7.8. Profitabilität YTD (kumulierte Profitabilität) über (Reihe (1+Tagesprofitabilität))-1
    yield_excl_div_year_df = yield_year_from_values_day(yield_excl_div_day_df, values_day_df, logger)
    if yield_excl_div_year_df is not None:
        export_2D_df_to_excel_format(yield_excl_div_year_df, (settings or {}).get("Export", {}).get("yield_excl_div_year_to_excel", {}), logger)
        logger.info("Kumulative jährliche Rendite erfolgreich berechnet und exportiert.")
    else:
        logger.warning("Kumulative jährliche Rendite (Yield) konnte nicht erstellt werden.")

    # 7.9. Renditekomponenten detailliert (Kursrendite, Dividenden, Gebühren, Steuern, Gesamtrendite)
    try:
        logger.info("Berechne detaillierte Renditekomponenten...")
        yield_components_day_df = yield_components_day(
            gains_losses_before_fees_taxes_day_df,
            fees_df,
            taxes_df,
            interest_dividends_df,
            values_day_df,
            transaction_value_at_price_day_df,
            logger
        )

        # Exportiere jede Komponente als separate Datei (analog zu bestehenden Exporten)
        component_columns = ['yield_price', 'yield_dividends', 'yield_fees', 'yield_taxes', 'yield_total']
        export_keys = ['yield_price_day_to_excel', 'yield_dividends_day_to_excel',
                       'yield_fees_day_to_excel', 'yield_taxes_day_to_excel', 'yield_total_day_to_excel']

        for col, key in zip(component_columns, export_keys):
            export_2D_df_to_excel_format(
                yield_components_day_df[[col]],
                (settings or {}).get("Export", {}).get(key, {}),
                logger
            )

        logger.info("Detaillierte Renditekomponenten (Tag) erfolgreich berechnet und als 5 separate Dateien exportiert.")
    except Exception as e:
        logger.error(f"Fehler bei Berechnung der Renditekomponenten (Tag): {e}")
        logger.error("Programmausführung wird fortgesetzt, aber yield_components_day_df ist nicht verfügbar.")
        yield_components_day_df = None

    # 7.10. Jährliche Renditekomponenten (kumuliert)
    if yield_components_day_df is not None:
        try:
            logger.info("Berechne jährliche Renditekomponenten...")
            yield_components_year_df = yield_components_year(
                yield_components_day_df,
                values_day_df,
                logger
            )

            # Exportiere jede Komponente als separate Datei
            component_columns = ['annual_yield_price', 'annual_yield_dividends', 'annual_yield_fees',
                                'annual_yield_taxes', 'annual_yield_total']
            export_keys = ['yield_price_year_to_excel', 'yield_dividends_year_to_excel',
                          'yield_fees_year_to_excel', 'yield_taxes_year_to_excel', 'yield_total_year_to_excel']

            for col, key in zip(component_columns, export_keys):
                export_2D_df_to_excel_format(
                    yield_components_year_df[[col]],
                    (settings or {}).get("Export", {}).get(key, {}),
                    logger
                )

            logger.info("Jährliche Renditekomponenten erfolgreich berechnet und als 5 separate Dateien exportiert.")
        except Exception as e:
            logger.error(f"Fehler bei Berechnung der jährlichen Renditekomponenten: {e}")
            logger.error("Programmausführung wird fortgesetzt.")
    else:
        logger.warning("Jährliche Renditekomponenten können nicht berechnet werden (yield_components_day_df fehlt).")

    # 7.11. Jahres-Profitabilitäts-Tabelle (Jahr als Zeilen, WKN mit days/yield als Spalten)
    if yield_components_day_df is not None and values_day_df is not None:
        try:
            logger.info("Erstelle Jahres-Profitabilitäts-Tabelle...")
            profitability_year_df = profitability_year_table(
                yield_components_day_df,
                values_day_df,
                logger
            )

            # Export mit Formatierung
            # NOTE: Cannot use export_2D_df_to_excel_format because it creates duplicate column names
            # when unstacking (year, wkn) index with (days, yield) columns. Using custom openpyxl export.
            export_config = (settings or {}).get("Export", {}).get('profitability_year_to_excel', {})
            if export_config.get('enabled', False):
                filename = export_config.get('filename', '')
                if filename:
                    # Direkter Export mit openpyxl (mit Formatierung)
                    try:
                        from openpyxl import Workbook
                        from openpyxl.worksheet.table import Table, TableStyleInfo

                        wb = Workbook()
                        ws = wb.active
                        ws.title = 'Sheet1'

                        # Schreibe Header
                        ws.cell(row=1, column=1, value=profitability_year_df.index.name or 'year')
                        for col_idx, col_name in enumerate(profitability_year_df.columns, start=2):
                            ws.cell(row=1, column=col_idx, value=col_name)

                        # Schreibe Daten mit Formatierung
                        for row_idx, (year, row_data) in enumerate(profitability_year_df.iterrows(), start=2):
                            # Jahr-Spalte (Index)
                            cell = ws.cell(row=row_idx, column=1, value=year)
                            cell.number_format = '0'

                            # Daten-Spalten
                            for col_idx, (col_name, value) in enumerate(row_data.items(), start=2):
                                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                                if '_days' in col_name:
                                    # Tage: Integer
                                    cell.number_format = '#,##0'
                                elif '_yield' in col_name:
                                    # Yield: Prozent mit 2 Nachkommastellen
                                    cell.number_format = '0.00%'

                        # Konvertiere zu Excel-Tabelle
                        max_col_letter = ws.cell(row=1, column=len(profitability_year_df.columns) + 1).column_letter
                        table_range = f"A1:{max_col_letter}{len(profitability_year_df) + 1}"

                        tab = Table(displayName="ProfitabilityYear", ref=table_range)
                        style = TableStyleInfo(
                            name="TableStyleMedium9",
                            showFirstColumn=False,
                            showLastColumn=False,
                            showRowStripes=True,
                            showColumnStripes=False
                        )
                        tab.tableStyleInfo = style
                        ws.add_table(tab)

                        # Spaltenbreiten und Header-Formatierung
                        from openpyxl.styles import Alignment

                        ws.column_dimensions['A'].width = 8
                        # Jahr-Header: 90° Ausrichtung
                        ws.cell(row=1, column=1).alignment = Alignment(textRotation=90, horizontal='center', vertical='bottom')

                        for col_idx, col_name in enumerate(profitability_year_df.columns, start=2):
                            col_letter = ws.cell(row=1, column=col_idx).column_letter

                            # Spaltenbreiten
                            if '_days' in col_name:
                                ws.column_dimensions[col_letter].width = 4.57
                            elif '_yield' in col_name:
                                ws.column_dimensions[col_letter].width = 7.71

                            # Header: 90° Textausrichtung
                            ws.cell(row=1, column=col_idx).alignment = Alignment(textRotation=90, horizontal='center', vertical='bottom')

                        wb.save(filename)
                        logger.info(f"DataFrame erfolgreich in '{filename}' exportiert.")
                        logger.info(f"Profitability-Tabelle formatiert und als Excel-Tabelle exportiert.")
                    except Exception as export_error:
                        logger.error(f"Export fehlgeschlagen: {export_error}")
                else:
                    logger.warning("Kein Dateiname für profitability_year_to_excel konfiguriert.")
            else:
                logger.info("Export von profitability_year ist deaktiviert.")

            logger.info("Jahres-Profitabilitäts-Tabelle erfolgreich erstellt und exportiert.")
        except Exception as e:
            logger.error(f"Fehler bei Erstellung der Jahres-Profitabilitäts-Tabelle: {e}")
            logger.error("Programmausführung wird fortgesetzt.")
    else:
        logger.warning("Jahres-Profitabilitäts-Tabelle kann nicht erstellt werden (yield_components_day_df oder values_day_df fehlt).")


    # 8. invest (Einschuss/Entnahme)
    bookings_filename = (settings or {}).get('Files', {}).get('bookings', '')
    invest_day_df = invest_day(bookings_filename, start_date, end_date, logger)
    if invest_day_df is not None:
        logger.info("Investitions-Daten auf Tagesbasis erfolgreich erstellt.")
    else:
        logger.error("Fehler beim Erstellen der Investitions-Daten auf Tagesbasis.")    
    

    # 9. Exportiert die Wert Aufteilung des Portfolios als absolute Werte und prozentuale Aufteilung 
    export_portfolio_analysis(values_day_df, instruments_type_df, instruments_region_df, logger)

    # 10. Abweichung von Zielportfolio (nach Einzel Instrumenten) und Rebalancing Vorschlag
    # Benötigte Daten: values_month_df, instruments_type_df
    # Benötigte externe Datei: provisions.xlsx, instruments.xlsx

    # 10.1. Lese die Rückstellungen (provisions)
    provisions_month_df = provisions_month_import_and_process(values_month_df, settings, logger)
    if provisions_month_df is None:
        sys.exit(1)  # Programm beenden, wenn die Provisionsdaten nicht erfolgreich verarbeitet wurden
    if (settings or {}).get('Export', {}).get('provisions_month_to_excel', {}):
        if (settings or {}).get("Export", {}).get("provisions_month_to_excel", {}).get("enabled", False):
            export_df_to_excel(provisions_month_df, (settings or {}).get("Export", {}).get("provisions_month_to_excel", {}).get("filename", ""), logger)

    # 10.2. Anpassung von values_month_df basierend auf provisions_month_df
    #
    # ich habe die Funktion deaktiviert, weil es bei der Subtraktion der Provisions von cash, eigentlich cash+cm+ftd+crypt für das Kriterium herangezogen 
    # werden müßte-
    # ich habe keine Verwendung des Dataframe "values_month_after_provisions_df" gefunden
    # ich bezalte erste einmal die Berechnung als Kommentar, falls ich sie wieder aktivieren muss
    #
    #values_month_after_provisions_df = values_month_adjust_for_provisions(values_month_df, provisions_month_df, logfile, screen=screen)
    #screen_and_log("Info: Werte (values) nach Anpassung durch Provisions erfolgreich erstellt.", logfile, screen=screen)
    #export_2D_df_to_excel_format(values_month_after_provisions_df, settings["Export"]["values_month_after_provisions_to_excel"], logfile, screen=False)

    # 10.3. Lese Zusammensetzung Zielportfolio
    target_shares_df=target_shares_import_and_process((settings or {}).get('Files', {}).get('instruments', ''), logger)
    if target_shares_df is not None:
        logger.info("Zielportfolio eingelesen und geprüft")

        # 10.4. Abweichung ermitteln auf Accountbasis
    
    
        # 10.5. Rebalancing Vorschlag
        buy_sell_df=values_vs_target(values_month_df, target_shares_df, prices_df, logger)
        if (settings or {}).get("Export", {}).get("buy_sell_to_excel", {}).get("enabled", False):
            export_df_to_excel(buy_sell_df, (settings or {}).get("Export", {}).get("buy_sell_to_excel", {}).get("filename", ""), logger)


    else:
        logger.warning("Zielportfolio eingelesen, aber Summe ergibt nicht 100%. Keine weitere Bearbeitung")


    # 10.4. Portfolio Zusammensetzung nach Korrektur des Cash Wertes um Rückstellungen (Provisions)
    values_type_month_df = values_type_month(values_month_df, instruments_type_df)
    
    values_type_month_after_provisions_df = values_type_month_after_provisions(values_type_month_df, provisions_month_df, logger)
    export_2D_df_to_excel_format(values_type_month_after_provisions_df, (settings or {}).get("Export", {}).get("values_type_month_after_provisions_to_excel", {}), logger)

    values_type_month_after_provisions_percentage_df = df_transform_each_line_to_percentage(values_type_month_after_provisions_df)
    export_2D_df_to_excel_format(values_type_month_after_provisions_percentage_df, (settings or {}).get("Export", {}).get("values_type_month_after_provisions_percentage_to_excel", {}), logger)

    # 11. Instrument Profitabilität
    # 11.1. Zinsen und Dividenden aus Bookings ermitteln 
    # 11.2. Gewinne pro Instrument pro Tag = Kursgewinne + Zins/Dividenden
    # 11.3. Profitabilität pro Tag = Gewinn / Eingesetztes Kaptial 
    # 11.4. Profitabilität YTD und annualisiert
    #       Ansatz 1: (1+rendite Tag1) * (1+rendite Tag2) * ... * (1+rendite letzter Tag) - 1   
    #       Ansatz 2: ln(1+rendite Tag1) + ln (1+rendite Tag2) + ... + ln(1+rendite letzter Tag) => exp(Summe)-1

    # 12. Ermittelung von Gebühren, Zinsen, Steuern aus Buchungen bookings.xls
    # ich sollte mir noch einmal die Struktur überlegen, wie ich Buchungen darstelle. 
    # in delta ist der nett cash-Buchungsbetrag, es werden gebühren, steuern nicht als buchung sondern als intrinsisch enthaltener wert dargetsellt, 
    # der gewinn für das intrument wird im Moment als Netto gewinn nach Abzug von Steuern und Gebühren angezeigt.





    # 19. Overview (Bericht)
    export_overview(values_day_df, unrealized_gains_losses_day_df, invest_day_df, logger)


    # 20. Depotauszug pro Bank (in einem File) um Wert und Anzahl bzw. Cash mit Konto-/Depot-Auszug zu vergleichen
    shares_month_banks_df = df_to_eom(shares_day_banks_df)
    if (settings or {}).get("Export", {}).get("depotauszug_to_excel", {}).get("enabled", False):
        export_bank_analysis_to_excel(shares_month_banks_df, values_month_banks_df, (settings or {}).get("Export", {}).get("depotauszug_to_excel", {}).get("filename", ""), logger)

    # 21. Depots für Finance analysieren und für Import bereitstellen

    depots_fuer_finance_df = depots_fuer_finance(values_month_banks_df, logger)
    if depots_fuer_finance_df is not None:
        if (settings or {}).get("Export", {}).get("depots_fuer_finance_to_excel", {}).get("enabled", False):
            export_df_to_excel(depots_fuer_finance_df, (settings or {}).get("Export", {}).get("depots_fuer_finance_to_excel", {}).get("filename", ""), logger)
        if (settings or {}).get("Export", {}).get("depots_fuer_finance_to_excel", {}).get("enabled", False):
            format_excel_as_table_with_freeze((settings or {}).get("Export", {}).get("depots_fuer_finance_to_excel", {}).get("filename", ""), logger)
            format_excel_columns((settings or {}).get("Export", {}).get("depots_fuer_finance_to_excel", {}).get("filename", ""), ["DD.MM.YY","#,##0.00"], logger, [12,12])


        