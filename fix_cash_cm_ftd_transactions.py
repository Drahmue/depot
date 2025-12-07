import pandas as pd
import numpy as np

# Lese bookings.xlsx
file_bookings = r'\\WIN-H7BKO5H0RMC\Dataserver\Dummy\Finance_Input\bookings.xlsx'
df_bookings = pd.read_excel(file_bookings)
df_bookings['date'] = pd.to_datetime(df_bookings['date'])

print('=' * 80)
print('KORREKTUREN FÜR CASH, CM, FTD - TRANSACTION_VALUE_AT_PRICE')
print('=' * 80)

# Instrumente die wir prüfen
instruments = ['cash', 'cm', 'ftd']

corrections = []

for instr in instruments:
    # Filter für dieses Instrument
    mask = df_bookings['wkn'].str.lower() == instr.lower()
    df_instr = df_bookings[mask].copy()

    # Finde alle mit fehlendem transaction_value_at_price
    # (NaN oder 0, wenn es keine Zinsbuchung ist)
    missing_mask = (
        (df_instr['transaction_value_at_price'].isna()) |
        (df_instr['transaction_value_at_price'] == 0)
    ) & (
        # Aber nur wenn interest_dividends AUCH leer ist (sonst ist es eine Zinsbuchung)
        (df_instr['interest_dividends'].isna()) |
        (df_instr['interest_dividends'] == 0)
    )

    df_missing = df_instr[missing_mask]

    if len(df_missing) > 0:
        print(f'\n{instr.upper()}: {len(df_missing)} Buchungen ohne transaction_value_at_price')

        for idx, row in df_missing.iterrows():
            delta = row['delta']

            # Für cash/cm/ftd: Preis ist immer 1.00 EUR
            # transaction_value_at_price = -(delta × 1.00) = -delta
            transaction_value = -delta

            corrections.append({
                'Excel_Row': idx + 2,  # +2 für Header und 0-basiertem Index
                'Date': row['date'],
                'WKN': row['wkn'],
                'Bank': row['bank'],
                'Delta': delta,
                'Current_Transaction': row['transaction_value_at_price'],
                'New_Transaction': transaction_value,
                'Interest_Dividends': row['interest_dividends']
            })

# Sortiere nach Datum und WKN
corrections = sorted(corrections, key=lambda x: (x['Date'], x['WKN']))

print(f'\n\n{"="*80}')
print(f'GESAMTÜBERSICHT: {len(corrections)} Korrekturen gefunden')
print(f'{"="*80}')

if len(corrections) > 0:
    # Gruppiere nach WKN
    by_wkn = {}
    for corr in corrections:
        wkn = corr['WKN'].upper()
        if wkn not in by_wkn:
            by_wkn[wkn] = []
        by_wkn[wkn].append(corr)

    print(f'\n{"WKN":6} | {"Anzahl":>8}')
    print('-' * 20)
    for wkn in sorted(by_wkn.keys()):
        print(f'{wkn:6} | {len(by_wkn[wkn]):8d}')

    # Detaillierte Liste
    print(f'\n\n{"="*80}')
    print('DETAILLIERTE KORREKTUREN')
    print(f'{"="*80}')
    print(f'{"Zeile":>6} | {"Datum":12} | {"WKN":6} | {"Bank":10} | {"Delta":>12} | {"Aktuell":>12} | {"NEU":>12}')
    print('-' * 80)

    for corr in corrections[:50]:  # Zeige erste 50
        current = corr['Current_Transaction']
        current_str = f'{current:.2f}' if pd.notna(current) else 'NaN'

        print(f'{corr["Excel_Row"]:6d} | {corr["Date"].strftime("%Y-%m-%d"):12} | '
              f'{corr["WKN"]:6} | {corr["Bank"]:10} | {corr["Delta"]:12.2f} | '
              f'{current_str:>12} | {corr["New_Transaction"]:12.2f}')

    if len(corrections) > 50:
        print(f'\n... und {len(corrections) - 50} weitere')

    # Export als CSV
    df_corrections = pd.DataFrame(corrections)
    output_file = 'cash_cm_ftd_corrections.csv'
    df_corrections.to_csv(output_file, index=False, sep=';', decimal=',', encoding='utf-8-sig')

    print(f'\n\n{"="*80}')
    print('EXPORT')
    print(f'{"="*80}')
    print(f'Datei gespeichert: {output_file}')
    print(f'Format: CSV mit Semikolon-Trennung')

    # Export als Excel für einfacheres Kopieren
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Korrekturen'

    # Header
    headers = ['Excel Zeile', 'Datum', 'WKN', 'Bank', 'Delta',
               'Aktuell', 'NEU', 'Interest_Dividends']

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Daten
    for row_num, corr in enumerate(corrections, 2):
        ws.cell(row=row_num, column=1, value=corr['Excel_Row'])
        ws.cell(row=row_num, column=2, value=corr['Date'].strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=3, value=corr['WKN'])
        ws.cell(row=row_num, column=4, value=corr['Bank'])
        ws.cell(row=row_num, column=5, value=corr['Delta'])

        current = corr['Current_Transaction']
        ws.cell(row=row_num, column=6, value=current if pd.notna(current) else 0)

        cell_new = ws.cell(row=row_num, column=7, value=corr['New_Transaction'])
        # NEU-Spalte hervorheben
        cell_new.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        cell_new.font = Font(bold=True)

        interest = corr['Interest_Dividends']
        ws.cell(row=row_num, column=8, value=interest if pd.notna(interest) else 0)

    # Spaltenbreiten
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 18

    excel_file = 'cash_cm_ftd_corrections.xlsx'
    wb.save(excel_file)

    print(f'Excel-Datei erstellt: {excel_file}')
    print('\nDie Spalte "NEU" (gelb markiert) enthält die Werte für transaction_value_at_price')

else:
    print('\nKeine Korrekturen nötig!')

print(f'\n{"="*80}')
print('ANLEITUNG')
print(f'{"="*80}')
print('''
1. Öffne cash_cm_ftd_corrections.xlsx
2. Die Spalte "NEU" (gelb markiert) enthält die korrekten Werte
3. Öffne bookings.xlsx
4. Für jede Zeile in cash_cm_ftd_corrections.xlsx:
   - Gehe zur entsprechenden Zeile in bookings.xlsx (siehe "Excel Zeile")
   - Trage den Wert aus "NEU" in die Spalte "transaction_value_at_price" ein
5. Speichere bookings.xlsx
6. Führe depot.py erneut aus

WICHTIG:
- Die Formel ist: transaction_value_at_price = -delta
- Bei Einzahlung (+delta): Negativer Wert (Geld fließt ab)
- Bei Auszahlung (-delta): Positiver Wert (Geld fließt zurück)
''')
print(f'{"="*80}')
